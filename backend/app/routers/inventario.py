from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import Optional
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.database import get_db
from app.core.security import get_current_user

router = APIRouter(prefix="/api/inventario", tags=["inventario"])


def get_proyectos_permitidos(current_user, db: Session) -> list[int]:
    """Retorna los IDs de proyectos que el usuario puede ver."""
    if current_user.nivel >= 3:  # GERENTE o ADMIN ven todos
        result = db.execute(text("SELECT id FROM proyectos WHERE activo=TRUE")).fetchall()
        return [r[0] for r in result]
    else:
        result = db.execute(
            text("SELECT proyecto_id FROM usuario_proyectos WHERE usuario_id=:uid"),
            {"uid": str(current_user.id)}
        ).fetchall()
        return [r[0] for r in result]


@router.get("/resumen")
async def get_resumen(
    año: Optional[int] = Query(None),
    mes: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """
    KPIs consolidados de inventario.
    Con año/mes: lotes con fecha_venta > cutoff_date se tratan como DISPONIBLES.
    Sin período: estado actual.
    """
    proyectos_ids = get_proyectos_permitidos(current_user, db)
    if not proyectos_ids:
        return {"proyectos": [], "totales": {}}

    ids_str = ",".join(str(i) for i in proyectos_ids)

    # Define cutoff BEFORE any query uses it
    import calendar as _cal
    if año and mes:
        _last = _cal.monthrange(año, mes)[1]
        cutoff = f"{año}-{mes:02d}-{_last}"
    else:
        cutoff = None

    # Per-project breakdown with period awareness
    if año and mes:
        resumen = db.execute(text(f"""
            SELECT
                p.id AS proyecto_id,
                p.nombre_proyecto,
                p.nombre_sociedad,
                p.empresa_sap,
                COUNT(l.id) AS total_lotes,
                COUNT(l.id) FILTER (WHERE
                    CASE WHEN l.fecha_venta IS NOT NULL AND l.fecha_venta > CAST(:cutoff AS DATE)
                    THEN 'DISPONIBLE' ELSE l.estatus END = 'DISPONIBLE') AS disponibles,
                COUNT(l.id) FILTER (WHERE
                    CASE WHEN l.fecha_venta IS NOT NULL AND l.fecha_venta > CAST(:cutoff AS DATE)
                    THEN 'DISPONIBLE' ELSE l.estatus END IN ('VENTA','RESERVADO')) AS vendidos_reservados,
                COUNT(l.id) FILTER (WHERE
                    CASE WHEN l.fecha_venta IS NOT NULL AND l.fecha_venta > CAST(:cutoff AS DATE)
                    THEN 'DISPONIBLE' ELSE l.estatus END = 'BLOQUEADO') AS bloqueados,
                COUNT(l.id) FILTER (WHERE
                    CASE WHEN l.fecha_venta IS NOT NULL AND l.fecha_venta > CAST(:cutoff AS DATE)
                    THEN 'DISPONIBLE' ELSE l.estatus END = 'CANJE') AS canjes,
                COALESCE(SUM(l.precio_final) FILTER (WHERE
                    CASE WHEN l.fecha_venta IS NOT NULL AND l.fecha_venta > CAST(:cutoff AS DATE)
                    THEN 'DISPONIBLE' ELSE l.estatus END = 'DISPONIBLE'), 0) AS valor_disponible,
                COALESCE(SUM(l.precio_final) FILTER (WHERE
                    CASE WHEN l.fecha_venta IS NOT NULL AND l.fecha_venta > CAST(:cutoff AS DATE)
                    THEN 'DISPONIBLE' ELSE l.estatus END IN ('VENTA','RESERVADO')), 0) AS valor_vendido,
                COALESCE(SUM(l.precio_final) FILTER (WHERE
                    CASE WHEN l.fecha_venta IS NOT NULL AND l.fecha_venta > CAST(:cutoff AS DATE)
                    THEN 'DISPONIBLE' ELSE l.estatus END = 'BLOQUEADO'), 0) AS valor_bloqueado,
                COALESCE(SUM(l.precio_final) FILTER (WHERE
                    CASE WHEN l.fecha_venta IS NOT NULL AND l.fecha_venta > CAST(:cutoff AS DATE)
                    THEN 'DISPONIBLE' ELSE l.estatus END = 'CANJE'), 0) AS valor_canjes,
                COALESCE(SUM(l.precio_final), 0) AS valor_total,
                ROUND(COUNT(l.id) FILTER (WHERE
                    CASE WHEN l.fecha_venta IS NOT NULL AND l.fecha_venta > CAST(:cutoff AS DATE)
                    THEN 'DISPONIBLE' ELSE l.estatus END IN ('VENTA','RESERVADO'))::NUMERIC
                    / NULLIF(COUNT(l.id) FILTER (WHERE
                    CASE WHEN l.fecha_venta IS NOT NULL AND l.fecha_venta > CAST(:cutoff AS DATE)
                    THEN 'DISPONIBLE' ELSE l.estatus END NOT IN ('BLOQUEADO','CANJE')), 0) * 100, 2) AS porcentaje_absorcion
            FROM lotes l
            JOIN proyectos p ON p.id = l.proyecto_id
            WHERE l.proyecto_id IN ({ids_str})
            GROUP BY p.id, p.nombre_proyecto, p.nombre_sociedad, p.empresa_sap
            ORDER BY p.nombre_proyecto
        """), {"cutoff": cutoff}).fetchall()
    else:
        resumen = db.execute(text(f"""
            SELECT * FROM v_resumen_inventario
            WHERE proyecto_id IN ({ids_str})
            ORDER BY nombre_proyecto
        """)).fetchall()

    # Totales consolidados — cutoff already defined above
    if año and mes:
        period_params = {"cutoff": cutoff}
        # Effective estatus: if fecha_venta > cutoff → DISPONIBLE, else real estatus
        estatus_expr = """
            CASE
                WHEN l.fecha_venta IS NOT NULL AND l.fecha_venta > CAST(:cutoff AS DATE)
                THEN 'DISPONIBLE'
                ELSE l.estatus
            END"""
    else:
        period_params = {}
        estatus_expr = "l.estatus"

    totales = db.execute(text(f"""
        SELECT
            COUNT(l.id)                                                                        AS total_lotes,
            COUNT(l.id) FILTER (WHERE ({estatus_expr})='DISPONIBLE')                          AS disponibles,
            COUNT(l.id) FILTER (WHERE ({estatus_expr}) IN ('VENTA','RESERVADO'))              AS vendidos,
            COUNT(l.id) FILTER (WHERE ({estatus_expr})='BLOQUEADO')                           AS bloqueados,
            COUNT(l.id) FILTER (WHERE ({estatus_expr})='CANJE')                               AS canjes,
            COALESCE(SUM(l.precio_final) FILTER (WHERE ({estatus_expr})='DISPONIBLE'), 0)             AS valor_disponible,
            COALESCE(SUM(l.precio_final) FILTER (WHERE ({estatus_expr}) IN ('VENTA','RESERVADO')), 0) AS valor_vendido,
            COALESCE(SUM(l.precio_final) FILTER (WHERE ({estatus_expr})='BLOQUEADO'), 0)              AS valor_bloqueado,
            COALESCE(SUM(l.precio_final) FILTER (WHERE ({estatus_expr})='CANJE'), 0)                  AS valor_canjes,
            COALESCE(SUM(l.precio_final), 0)                                                  AS valor_total,
            ROUND(COUNT(l.id) FILTER (WHERE ({estatus_expr}) IN ('VENTA','RESERVADO'))::NUMERIC
                  / NULLIF(COUNT(l.id) FILTER (WHERE ({estatus_expr}) NOT IN ('BLOQUEADO','CANJE')),0)*100, 2) AS pct_absorcion
        FROM lotes l
        WHERE l.proyecto_id IN ({ids_str})
    """), period_params).fetchone()

    # Última sincronización
    last_sync = db.execute(text("""
        SELECT inicio, fin, estado, registros_actualizados+registros_insertados AS procesados
        FROM sync_log WHERE archivo='INVENTARIO' AND estado='EXITOSO'
        ORDER BY fin DESC LIMIT 1
    """)).fetchone()

    return {
        "proyectos": [dict(r._mapping) for r in resumen],
        "totales": dict(totales._mapping) if totales else {},
        "ultima_sincronizacion": dict(last_sync._mapping) if last_sync else None
    }


@router.get("/proyectos")
async def get_proyectos(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    proyectos_ids = get_proyectos_permitidos(current_user, db)
    if not proyectos_ids:
        return []
    ids_str = ",".join(str(i) for i in proyectos_ids)
    result = db.execute(text(f"""
        SELECT id, empresa_sap, nombre_sociedad, nombre_proyecto
        FROM proyectos WHERE id IN ({ids_str}) AND activo=TRUE
        ORDER BY nombre_proyecto
    """)).fetchall()
    return [dict(r._mapping) for r in result]


@router.get("/lotes")
async def get_lotes(
    proyecto_id: Optional[int] = Query(None),
    estatus: Optional[str] = Query(None),
    manzana: Optional[str] = Query(None),
    buscar: Optional[str] = Query(None),
    forma_pago: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Lista de lotes con filtros y paginación."""
    proyectos_ids = get_proyectos_permitidos(current_user, db)
    if not proyectos_ids:
        return {"lotes": [], "total": 0, "page": page, "pages": 0, "stats": {}}

    ids_str = ",".join(str(i) for i in proyectos_ids)
    conditions = [f"l.proyecto_id IN ({ids_str})"]
    params = {}

    if proyecto_id:
        if proyecto_id not in proyectos_ids:
            raise HTTPException(status_code=403, detail="Sin acceso a ese proyecto")
        conditions.append("l.proyecto_id = :proyecto_id")
        params["proyecto_id"] = proyecto_id

    if manzana:
        conditions.append("l.manzana ILIKE :manzana")
        params["manzana"] = f"%{manzana}%"

    if buscar:
        conditions.append("(l.card_name ILIKE :buscar OR l.unidad_key ILIKE :buscar OR l.manzana ILIKE :buscar)")
        params["buscar"] = f"%{buscar}%"

    if forma_pago:
        conditions.append("l.forma_pago = :forma_pago")
        params["forma_pago"] = forma_pago

    # WHERE sin filtro de estatus → para calcular stats por estatus siempre completos
    where_base = " AND ".join(conditions)

    # Stats por estatus (siempre sobre el filtro base, sin estatus)
    stats_rows = db.execute(text(f"""
        SELECT
            estatus,
            COUNT(*) as cantidad,
            COALESCE(SUM(precio_final), 0) as valor_total
        FROM lotes l
        WHERE {where_base}
        GROUP BY estatus
    """), params).fetchall()

    stats_dict = {}
    for s in stats_rows:
        stats_dict[s.estatus] = {
            "cantidad": s.cantidad,
            "valor_total": float(s.valor_total)
        }

    # Ahora aplicar filtro de estatus para la lista de lotes
    conditions_with_estatus = list(conditions)
    params_with_estatus = dict(params)

    if estatus:
        # VENTA incluye RESERVADO (mismo concepto en SAP)
        if estatus.upper() == "VENTA":
            conditions_with_estatus.append("l.estatus IN ('VENTA','RESERVADO')")
        else:
            conditions_with_estatus.append("l.estatus = :estatus")
            params_with_estatus["estatus"] = estatus.upper()

    where = " AND ".join(conditions_with_estatus)
    offset = (page - 1) * page_size

    total = db.execute(text(f"SELECT COUNT(*) FROM lotes l WHERE {where}"), params_with_estatus).scalar()

    lotes = db.execute(text(f"""
        SELECT
            l.id, l.unidad_key, l.manzana, l.metraje_inventario,
            l.estatus, l.precio_final, l.precio_sin_descuento,
            l.card_name, l.vendedor, l.fecha_venta,
            l.forma_pago, l.plazo, l.saldo_cliente,
            l.pagado_capital, l.pendiente_capital,
            l.cuotas_pagadas, l.cuotas_pendientes,
            l.es_esquina, l.cuota_mantenimiento,
            p.nombre_proyecto, p.nombre_sociedad
        FROM lotes l
        JOIN proyectos p ON p.id = l.proyecto_id
        WHERE {where}
        ORDER BY l.manzana, l.unidad_key
        LIMIT :limit OFFSET :offset
    """), {**params_with_estatus, "limit": page_size, "offset": offset}).fetchall()

    return {
        "lotes": [dict(r._mapping) for r in lotes],
        "total": total,
        "page": page,
        "pages": -(-total // page_size),
        "stats": stats_dict
    }




@router.get("/lotes/exportar")
async def exportar_lotes(
    proyecto_id: Optional[int] = Query(None),
    estatus: Optional[str] = Query(None),
    manzana: Optional[str] = Query(None),
    buscar: Optional[str] = Query(None),
    forma_pago: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Exporta lotes a Excel con los mismos filtros de la vista."""
    proyectos_ids = get_proyectos_permitidos(current_user, db)
    if not proyectos_ids:
        raise HTTPException(status_code=403, detail="Sin acceso")

    ids_str = ",".join(str(i) for i in proyectos_ids)
    conditions = [f"l.proyecto_id IN ({ids_str})"]
    params = {}

    if proyecto_id:
        if proyecto_id not in proyectos_ids:
            raise HTTPException(status_code=403, detail="Sin acceso a ese proyecto")
        conditions.append("l.proyecto_id = :proyecto_id")
        params["proyecto_id"] = proyecto_id
    if manzana:
        conditions.append("l.manzana ILIKE :manzana")
        params["manzana"] = f"%{manzana}%"
    if buscar:
        conditions.append("(l.card_name ILIKE :buscar OR l.unidad_key ILIKE :buscar OR l.manzana ILIKE :buscar)")
        params["buscar"] = f"%{buscar}%"
    if estatus:
        conditions.append("l.estatus = :estatus")
        params["estatus"] = estatus
    if forma_pago:
        conditions.append("l.forma_pago = :forma_pago")
        params["forma_pago"] = forma_pago

    where = " AND ".join(conditions)
    rows = db.execute(text(f"""
        SELECT p.nombre_proyecto, p.nombre_sociedad,
               l.unidad_key, l.manzana, l.estatus,
               l.metraje_inventario, l.precio_final,
               l.forma_pago, l.plazo, l.plazo_raw,
               l.card_name AS cliente, l.vendedor,
               l.fecha_venta, l.fecha_solicitud_pcv,
               l.status_promesa_compraventa,
               l.pagado_capital, l.pendiente_capital,
               l.total_intereses, l.saldo_cliente
        FROM lotes l
        JOIN proyectos p ON p.id = l.proyecto_id
        WHERE {where}
        ORDER BY p.nombre_proyecto, l.manzana, l.unidad_key
    """), params).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Lotes"

    headers = ["Proyecto","Sociedad","Lote","Manzana","Estatus","M²","Precio Final",
               "Forma Pago","Plazo","Tipo Plazo","Cliente","Asesor",
               "Fecha Venta","Fecha PCV","Status PCV",
               "Pagado Capital","Pendiente Capital","Intereses","Saldo Cliente"]

    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(rows, 2):
        vals = [
            r.nombre_proyecto, r.nombre_sociedad,
            r.unidad_key, r.manzana, r.estatus,
            float(r.metraje_inventario) if r.metraje_inventario else None,
            float(r.precio_final) if r.precio_final else None,
            r.forma_pago, r.plazo, r.plazo_raw,
            r.cliente, r.vendedor,
            r.fecha_venta, r.fecha_solicitud_pcv, r.status_promesa_compraventa,
            float(r.pagado_capital) if r.pagado_capital else 0,
            float(r.pendiente_capital) if r.pendiente_capital else 0,
            float(r.total_intereses) if r.total_intereses else 0,
            float(r.saldo_cliente) if r.saldo_cliente else 0,
        ]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row_idx, column=col, value=v)

    # Auto width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "inventario_lotes.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/lotes/{lote_id}")
async def get_lote_detalle(
    lote_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Detalle completo de un lote."""
    proyectos_ids = get_proyectos_permitidos(current_user, db)
    ids_str = ",".join(str(i) for i in proyectos_ids)

    lote = db.execute(text(f"""
        SELECT l.*, p.nombre_proyecto, p.nombre_sociedad, p.empresa_sap
        FROM lotes l JOIN proyectos p ON p.id = l.proyecto_id
        WHERE l.id = :id AND l.proyecto_id IN ({ids_str})
    """), {"id": lote_id}).fetchone()

    if not lote:
        raise HTTPException(status_code=404, detail="Lote no encontrado")
    return dict(lote._mapping)


@router.get("/manzanas/{proyecto_id}")
async def get_manzanas(
    proyecto_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    proyectos_ids = get_proyectos_permitidos(current_user, db)
    if proyecto_id not in proyectos_ids:
        raise HTTPException(status_code=403)

    result = db.execute(text("""
        SELECT DISTINCT manzana, COUNT(*) as total,
               COUNT(*) FILTER (WHERE estatus='DISPONIBLE') as disponibles
        FROM lotes WHERE proyecto_id=:pid AND manzana IS NOT NULL
        GROUP BY manzana ORDER BY manzana
    """), {"pid": proyecto_id}).fetchall()

    return [dict(r._mapping) for r in result]