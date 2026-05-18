"""
proyecciones.py — Router FastAPI v3 — Proyecciones al Cierre
Módulos: Ingresos, Egresos Op (PPTO vs Ejecutado), Financieros (Préstamos+Intercompany), Tierra/Dividendos
"""

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import Optional
from datetime import date, datetime, timedelta
import json
import io

from app.database import get_db
from app.core.security import get_current_user

router = APIRouter(prefix="/api/proyecciones", tags=["proyecciones"])

# ─────────────────────────────────────────────
# DATOS ESTÁTICOS DESDE EXCEL (embebidos en código)
# ─────────────────────────────────────────────

# Presupuesto por sociedad (PPTO. PROYECTOS)
PPTO_PROYECTOS = {
    "Capipos":            {"urb": 2960886.97,  "adm": 5486340.625, "total": 8447227.595},
    "Ottavia":            {"urb": 32565496.94, "adm": 4435908.82,  "total": 37001405.76},
    "Vilet":              {"urb": 7776548.47,  "adm": 2529479.51,  "total": 10306027.98},
    "Tezzoli":            {"urb": 38712581.05, "adm": 5353552.61,  "total": 44066133.66},
    "Garbatella":         {"urb": 12140219.0,  "adm": 5234375.0,   "total": 17374594.0},
    "Urbiva":             {"urb": 18025593.4,  "adm": 4316677.41,  "total": 22342270.81},
    "Utilica":            {"urb": 15651579.86, "adm": 2740817.51,  "total": 18392397.37},
    "Corcolle":           {"urb": 2940898.26,  "adm": 681390.50,   "total": 3622288.76},
    "Leofreni":           {"urb": 4219549.67,  "adm": 977647.24,   "total": 5197196.91},
    "Gibraleon":          {"urb": 1150786.27,  "adm": 266631.07,   "total": 1417417.34},
    "Talocci":            {"urb": 3963819.39,  "adm": 918395.89,   "total": 4882215.28},
    "Eficiencia Urbana":  {"urb": 43682514.69, "adm": 12677589.09, "total": 56360103.77},
    "Ovest":              {"urb": 17956528.71, "adm": 2720307.85,  "total": 20676836.56},
    "Rossio":             {"urb": 14989329.93, "adm": 4209161.28,  "total": 19198491.21},
    "Servicios Generales":{"urb": 14416623.0,  "adm": 2836181.20,  "total": 17252804.20},
    "Frugalex":           {"urb": 8710467.37,  "adm": 3320576.61,  "total": 12031043.98},
}

# Mapeo empresa BD → clave PPTO
EMPRESA_TO_PPTO = {
    "Capipos, S.A.":               "Capipos",
    "Ottavia, S.A.":               "Ottavia",
    "Vilet, S.A.":                 "Vilet",
    "Tezzoli, S.A.":               "Tezzoli",
    "Garbatella, S.A.":            "Garbatella",
    "Urviba 2, S.A.":              "Urbiva",
    "Utilica, S.A.":               "Utilica",
    "Corcolle, S.A.":              "Corcolle",
    "Leofreni, S.A.":              "Leofreni",
    "Gibraleon, S.A.":             "Gibraleon",
    "Talocci, S.A.":               "Talocci",
    "Eficiencia Urbana, S. A.":    "Eficiencia Urbana",
    "Ovest, S.A.":                 "Ovest",
    "Rossio, S.A.":                "Rossio",
    "Servicios Generales CCC, S.A.":"Servicios Generales",
    "Frugalex, S.A.":              "Frugalex",
}

# Tablas de amortización de préstamos bancarios (desde PRESTAMOS BANCARIOS)
PRESTAMOS_AMORT = {
    "Eficiencia Urbana, S. A.": {
        "banco": "Banrural", "no_credito": "7991000667",
        "monto_original": 14000000, "tasa": 0.09,
        "cuotas": [
          {"op":1,"fecha":"2025-10-04","saldo_capital":14000000.0,"capital":0.0,"interes":57994.52,"cuota":57994.52},
          {"op":2,"fecha":"2025-11-04","saldo_capital":14000000.0,"capital":0.0,"interes":74909.59,"cuota":74909.59},
          {"op":3,"fecha":"2025-12-04","saldo_capital":14000000.0,"capital":72413.79,"interes":102526.03,"cuota":174939.82},
          {"op":4,"fecha":"2026-01-05","saldo_capital":13927586.21,"capital":72413.79,"interes":109154.66,"cuota":181568.45},
          {"op":5,"fecha":"2026-02-04","saldo_capital":13855172.42,"capital":72413.79,"interes":101916.39,"cuota":174330.18},
          {"op":6,"fecha":"2026-03-04","saldo_capital":13782758.63,"capital":72413.79,"interes":94123.33,"cuota":166537.12},
          {"op":7,"fecha":"2026-04-06","saldo_capital":13710344.84,"capital":132413.79,"interes":110887.48,"cuota":243301.27},
          {"op":8,"fecha":"2026-05-04","saldo_capital":13577931.05,"capital":132413.79,"interes":93743.52,"cuota":226157.31},
          {"op":9,"fecha":"2026-06-04","saldo_capital":13445517.26,"capital":132413.79,"interes":102775.32,"cuota":235189.11},
          {"op":10,"fecha":"2026-07-04","saldo_capital":13313103.47,"capital":132413.79,"interes":98480.49,"cuota":230894.28},
          {"op":11,"fecha":"2026-08-04","saldo_capital":13180689.68,"capital":132413.79,"interes":100751.03,"cuota":233164.82},
          {"op":12,"fecha":"2026-09-04","saldo_capital":13048275.89,"capital":132413.79,"interes":99738.88,"cuota":232152.67},
          {"op":13,"fecha":"2026-10-05","saldo_capital":12915862.1,"capital":132413.79,"interes":98726.73,"cuota":231140.52},
          {"op":14,"fecha":"2026-11-04","saldo_capital":12783448.31,"capital":132413.79,"interes":94562.49,"cuota":226976.28},
          {"op":15,"fecha":"2026-12-04","saldo_capital":12651034.52,"capital":132413.79,"interes":93583.0,"cuota":225996.79},
          {"op":16,"fecha":"2027-01-04","saldo_capital":12518620.73,"capital":132413.79,"interes":95690.28,"cuota":228104.07},
          {"op":17,"fecha":"2027-02-04","saldo_capital":12386206.94,"capital":132413.79,"interes":94678.13,"cuota":227091.92},
          {"op":18,"fecha":"2027-03-04","saldo_capital":12253793.15,"capital":132413.79,"interes":84601.53,"cuota":217015.32},
          {"op":19,"fecha":"2027-04-05","saldo_capital":12121379.36,"capital":132413.79,"interes":95642.66,"cuota":228056.45},
          {"op":20,"fecha":"2027-05-04","saldo_capital":11988965.57,"capital":132413.79,"interes":85729.32,"cuota":218143.11},
          {"op":21,"fecha":"2027-06-04","saldo_capital":11856551.78,"capital":132413.79,"interes":90629.53,"cuota":223043.32},
          {"op":22,"fecha":"2027-07-05","saldo_capital":11724137.99,"capital":132413.79,"interes":89617.38,"cuota":222031.17},
          {"op":23,"fecha":"2027-08-04","saldo_capital":11591724.2,"capital":132413.79,"interes":85747.0,"cuota":218160.79},
          {"op":24,"fecha":"2027-09-04","saldo_capital":11459310.41,"capital":132413.79,"interes":87593.09,"cuota":220006.88},
          {"op":25,"fecha":"2027-10-04","saldo_capital":11326896.62,"capital":132413.79,"interes":83788.0,"cuota":216201.79},
          {"op":26,"fecha":"2027-11-04","saldo_capital":11194482.83,"capital":132413.79,"interes":85568.79,"cuota":217982.58},
          {"op":27,"fecha":"2027-12-04","saldo_capital":11062069.04,"capital":132413.79,"interes":81829.0,"cuota":214242.79},
          {"op":28,"fecha":"2028-01-04","saldo_capital":10929655.25,"capital":132413.79,"interes":83544.49,"cuota":215958.28},
          {"op":29,"fecha":"2028-02-04","saldo_capital":10797241.46,"capital":132413.79,"interes":82532.34,"cuota":214946.13},
          {"op":30,"fecha":"2028-03-04","saldo_capital":10664827.67,"capital":132413.79,"interes":76260.82,"cuota":208674.61},
          {"op":31,"fecha":"2028-04-04","saldo_capital":10532413.88,"capital":132413.79,"interes":80508.04,"cuota":212921.83},
          {"op":32,"fecha":"2028-05-04","saldo_capital":10400000.09,"capital":132413.79,"interes":76931.51,"cuota":209345.30},
          {"op":33,"fecha":"2028-06-05","saldo_capital":10267586.3,"capital":132413.79,"interes":81015.48,"cuota":213429.27},
          {"op":34,"fecha":"2028-07-04","saldo_capital":10135172.51,"capital":132413.79,"interes":72473.43,"cuota":204887.22},
          {"op":35,"fecha":"2028-08-04","saldo_capital":10002758.72,"capital":132413.79,"interes":76459.44,"cuota":208873.23},
          {"op":36,"fecha":"2028-09-04","saldo_capital":9870344.93,"capital":132413.79,"interes":75447.29,"cuota":207861.08},
          {"op":37,"fecha":"2028-10-04","saldo_capital":9737931.14,"capital":132413.79,"interes":72034.01,"cuota":204447.80},
          {"op":38,"fecha":"2028-11-04","saldo_capital":9605517.35,"capital":132413.79,"interes":73423.0,"cuota":205836.79},
          {"op":39,"fecha":"2028-12-04","saldo_capital":9473103.56,"capital":132413.79,"interes":70075.01,"cuota":202488.80},
          {"op":40,"fecha":"2029-01-04","saldo_capital":9340689.77,"capital":132413.79,"interes":71398.7,"cuota":203812.49},
          {"op":41,"fecha":"2029-02-05","saldo_capital":9208275.98,"capital":132413.79,"interes":72657.08,"cuota":205070.87},
          {"op":42,"fecha":"2029-03-05","saldo_capital":9075862.19,"capital":132413.79,"interes":62660.75,"cuota":195074.54},
          {"op":43,"fecha":"2029-04-04","saldo_capital":8943448.4,"capital":132413.79,"interes":66157.02,"cuota":198570.81},
          {"op":44,"fecha":"2029-05-04","saldo_capital":8811034.61,"capital":132413.79,"interes":65177.52,"cuota":197591.31},
          {"op":45,"fecha":"2029-06-04","saldo_capital":8678620.82,"capital":132413.79,"interes":66337.95,"cuota":198751.74},
          {"op":46,"fecha":"2029-07-04","saldo_capital":8546207.03,"capital":132413.79,"interes":63218.52,"cuota":195632.31},
          {"op":47,"fecha":"2029-08-04","saldo_capital":8413793.24,"capital":132413.79,"interes":64313.65,"cuota":196727.44},
          {"op":48,"fecha":"2029-09-04","saldo_capital":8281379.45,"capital":132413.79,"interes":63301.5,"cuota":195715.29},
          {"op":49,"fecha":"2029-10-04","saldo_capital":8148965.66,"capital":132413.79,"interes":60280.02,"cuota":192693.81},
          {"op":50,"fecha":"2029-11-04","saldo_capital":8016551.87,"capital":132413.79,"interes":61274.65,"cuota":193688.44},
          {"op":51,"fecha":"2029-12-04","saldo_capital":7884138.08,"capital":132413.79,"interes":58318.16,"cuota":190731.95},
          {"op":52,"fecha":"2030-01-06","saldo_capital":7751724.29,"capital":132413.79,"interes":60873.41,"cuota":193287.20},
          {"op":53,"fecha":"2030-02-04","saldo_capital":7619310.5,"capital":132413.79,"interes":58251.73,"cuota":190665.52},
          {"op":54,"fecha":"2030-03-04","saldo_capital":7486896.71,"capital":132413.79,"interes":51727.45,"cuota":184141.24},
          {"op":55,"fecha":"2030-04-04","saldo_capital":7354482.92,"capital":132413.79,"interes":56222.95,"cuota":188636.74},
          {"op":56,"fecha":"2030-05-04","saldo_capital":7222069.13,"capital":132413.79,"interes":53439.51,"cuota":185853.30},
          {"op":57,"fecha":"2030-06-04","saldo_capital":7089655.34,"capital":132413.79,"interes":54192.72,"cuota":186606.51},
          {"op":58,"fecha":"2030-07-04","saldo_capital":6957241.55,"capital":132413.79,"interes":51474.78,"cuota":183888.57},
          {"op":59,"fecha":"2030-08-04","saldo_capital":6824827.76,"capital":132413.79,"interes":52155.95,"cuota":184569.74},
          {"op":60,"fecha":"2030-09-04","saldo_capital":6692413.97,"capital":6692413.97,"interes":49505.53,"cuota":6741919.5},
        ]
    },
    "Utilica, S.A.": {
        "banco": "Banrural", "no_credito": "7991000653",
        "monto_original": 5600000, "tasa": 0.09,
        "cuotas": [
          {"op":1,"fecha":"2025-10-06","saldo_capital":5600000.0,"capital":0.0,"interes":53852.05,"cuota":53852.05},
          {"op":2,"fecha":"2025-11-05","saldo_capital":5600000.0,"capital":0.0,"interes":41424.66,"cuota":41424.66},
          {"op":3,"fecha":"2025-12-05","saldo_capital":5600000.0,"capital":0.0,"interes":41424.66,"cuota":41424.66},
          {"op":4,"fecha":"2026-01-05","saldo_capital":5600000.0,"capital":93333.33,"interes":43517.81,"cuota":136851.14},
          {"op":5,"fecha":"2026-02-05","saldo_capital":5506666.67,"capital":93333.33,"interes":40758.47,"cuota":134091.80},
          {"op":6,"fecha":"2026-03-05","saldo_capital":5413333.34,"capital":93333.33,"interes":37063.28,"cuota":130396.61},
          {"op":7,"fecha":"2026-04-05","saldo_capital":5320000.01,"capital":93333.33,"interes":40661.64,"cuota":133994.97},
          {"op":8,"fecha":"2026-05-05","saldo_capital":5226666.68,"capital":93333.33,"interes":38683.01,"cuota":132016.34},
          {"op":9,"fecha":"2026-06-05","saldo_capital":5133333.35,"capital":93333.33,"interes":39239.73,"cuota":132573.06},
          {"op":10,"fecha":"2026-07-05","saldo_capital":5040000.02,"capital":93333.33,"interes":37295.89,"cuota":130629.22},
          {"op":11,"fecha":"2026-08-05","saldo_capital":4946666.69,"capital":93333.33,"interes":37817.81,"cuota":131151.14},
          {"op":12,"fecha":"2026-09-05","saldo_capital":4853333.36,"capital":93333.33,"interes":37092.47,"cuota":130425.80},
          {"op":13,"fecha":"2026-10-05","saldo_capital":4760000.03,"capital":93333.33,"interes":35314.52,"cuota":128647.85},
          {"op":14,"fecha":"2026-11-05","saldo_capital":4666666.70,"capital":93333.33,"interes":34545.21,"cuota":127878.54},
          {"op":15,"fecha":"2026-12-05","saldo_capital":4573333.37,"capital":93333.33,"interes":33819.18,"cuota":127152.51},
          {"op":16,"fecha":"2027-01-05","saldo_capital":4480000.04,"capital":93333.33,"interes":34238.36,"cuota":127571.69},
          {"op":17,"fecha":"2027-02-05","saldo_capital":4386666.71,"capital":93333.33,"interes":33513.01,"cuota":126846.34},
          {"op":18,"fecha":"2027-03-05","saldo_capital":4293333.38,"capital":93333.33,"interes":29634.25,"cuota":122967.58},
          {"op":19,"fecha":"2027-04-05","saldo_capital":4200000.05,"capital":93333.33,"interes":32100.0,"cuota":125433.33},
          {"op":20,"fecha":"2027-05-05","saldo_capital":4106666.72,"capital":93333.33,"interes":30393.15,"cuota":123726.48},
          {"op":21,"fecha":"2027-06-05","saldo_capital":4013333.39,"capital":93333.33,"interes":30660.27,"cuota":123993.60},
          {"op":22,"fecha":"2027-07-05","saldo_capital":3920000.06,"capital":93333.33,"interes":29952.33,"cuota":123285.66},
          {"op":23,"fecha":"2027-08-05","saldo_capital":3826666.73,"capital":93333.33,"interes":29260.27,"cuota":122593.60},
          {"op":24,"fecha":"2027-09-05","saldo_capital":3733333.40,"capital":93333.33,"interes":28511.51,"cuota":121844.84},
          {"op":25,"fecha":"2027-10-05","saldo_capital":3640000.07,"capital":93333.33,"interes":26958.90,"cuota":120292.23},
          {"op":26,"fecha":"2027-11-05","saldo_capital":3546666.74,"capital":93333.33,"interes":26219.18,"cuota":119552.51},
          {"op":27,"fecha":"2027-12-05","saldo_capital":3453333.41,"capital":93333.33,"interes":25616.44,"cuota":118949.77},
          {"op":28,"fecha":"2028-01-05","saldo_capital":3360000.08,"capital":93333.33,"interes":25679.45,"cuota":119012.78},
          {"op":29,"fecha":"2028-02-05","saldo_capital":3266666.75,"capital":93333.33,"interes":24972.60,"cuota":118305.93},
          {"op":30,"fecha":"2028-03-05","saldo_capital":3173333.42,"capital":93333.33,"interes":21923.29,"cuota":115256.62},
          {"op":31,"fecha":"2028-04-05","saldo_capital":3080000.09,"capital":93333.33,"interes":23527.40,"cuota":116860.73},
          {"op":32,"fecha":"2028-05-05","saldo_capital":2986666.76,"capital":93333.33,"interes":22116.44,"cuota":115449.77},
          {"op":33,"fecha":"2028-06-05","saldo_capital":2893333.43,"capital":93333.33,"interes":22118.63,"cuota":115451.96},
          {"op":34,"fecha":"2028-07-05","saldo_capital":2800000.10,"capital":93333.33,"interes":20720.55,"cuota":114053.88},
          {"op":35,"fecha":"2028-08-05","saldo_capital":2706666.77,"capital":93333.33,"interes":20684.93,"cuota":114018.26},
          {"op":36,"fecha":"2028-09-05","saldo_capital":2613333.44,"capital":93333.33,"interes":19979.45,"cuota":113312.78},
          {"op":37,"fecha":"2028-10-05","saldo_capital":2520000.11,"capital":93333.33,"interes":18655.07,"cuota":111988.40},
          {"op":38,"fecha":"2028-11-05","saldo_capital":2426666.78,"capital":93333.33,"interes":17945.21,"cuota":111278.54},
          {"op":39,"fecha":"2028-12-05","saldo_capital":2333333.45,"capital":93333.33,"interes":17268.49,"cuota":110601.82},
          {"op":40,"fecha":"2029-01-05","saldo_capital":2240000.12,"capital":93333.33,"interes":17123.29,"cuota":110456.62},
          {"op":41,"fecha":"2029-02-05","saldo_capital":2146666.79,"capital":93333.33,"interes":16431.51,"cuota":109764.84},
          {"op":42,"fecha":"2029-03-05","saldo_capital":2053333.46,"capital":93333.33,"interes":14175.34,"cuota":107508.67},
          {"op":43,"fecha":"2029-04-05","saldo_capital":1960000.13,"capital":93333.33,"interes":14980.27,"cuota":108313.60},
          {"op":44,"fecha":"2029-05-05","saldo_capital":1866666.80,"capital":93333.33,"interes":13819.18,"cuota":107152.51},
          {"op":45,"fecha":"2029-06-05","saldo_capital":1773333.47,"capital":93333.33,"interes":13551.78,"cuota":106885.11},
          {"op":46,"fecha":"2029-07-05","saldo_capital":1680000.14,"capital":93333.33,"interes":12427.40,"cuota":105760.73},
          {"op":47,"fecha":"2029-08-05","saldo_capital":1586666.81,"capital":93333.33,"interes":12127.40,"cuota":105460.73},
          {"op":48,"fecha":"2029-09-05","saldo_capital":1493333.48,"capital":93333.33,"interes":11416.44,"cuota":104749.77},
          {"op":49,"fecha":"2029-10-05","saldo_capital":1400000.15,"capital":93333.33,"interes":10367.12,"cuota":103700.45},
          {"op":50,"fecha":"2029-11-05","saldo_capital":1306666.82,"capital":93333.33,"interes":9672.88,"cuota":103006.21},
          {"op":51,"fecha":"2029-12-05","saldo_capital":1213333.49,"capital":93333.33,"interes":8979.45,"cuota":102312.78},
          {"op":52,"fecha":"2030-01-05","saldo_capital":1120000.16,"capital":93333.33,"interes":8562.74,"cuota":101896.07},
          {"op":53,"fecha":"2030-02-05","saldo_capital":1026666.83,"capital":93333.33,"interes":7843.15,"cuota":101176.48},
          {"op":54,"fecha":"2030-03-05","saldo_capital":933333.50,"capital":93333.33,"interes":6438.36,"cuota":99771.69},
          {"op":55,"fecha":"2030-04-05","saldo_capital":840000.17,"capital":93333.33,"interes":6424.93,"cuota":99758.26},
          {"op":56,"fecha":"2030-05-05","saldo_capital":746666.84,"capital":93333.33,"interes":5525.21,"cuota":98858.54},
          {"op":57,"fecha":"2030-06-05","saldo_capital":653333.51,"capital":93333.33,"interes":4993.15,"cuota":98326.48},
          {"op":58,"fecha":"2030-07-05","saldo_capital":560000.18,"capital":93333.33,"interes":4143.84,"cuota":97477.17},
          {"op":59,"fecha":"2030-08-05","saldo_capital":466666.85,"capital":93333.33,"interes":3567.12,"cuota":96900.45},
          {"op":60,"fecha":"2030-09-05","saldo_capital":156666.63,"capital":156666.63,"interes":1197.53,"cuota":157864.16},
        ]
    },
    "Servicios Generales CCC, S.A.": {
        "banco": "Banrural", "no_credito": "7951016725",
        "monto_original": 3000000, "tasa": 0.09,
        "cuotas": []  # Se cargan dinámicamente desde la BD - préstamo casi liquidado
    },
        "Capipos, S.A.": {
        "banco": "G&T Continental", "no_credito": "10-060179247",
        "monto_original": 11700000, "tasa": 0.0925,
        "cuotas": [
          {"op":1,"fecha":"2024-05-05","saldo_capital":11700000.0,"capital":0.0,"interes":14825.34,"cuota":14825.34},
          {"op":2,"fecha":"2024-06-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":3,"fecha":"2024-07-05","saldo_capital":11700000.0,"capital":0.0,"interes":88952.05,"cuota":88952.05},
          {"op":4,"fecha":"2024-08-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":5,"fecha":"2024-09-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":6,"fecha":"2024-10-05","saldo_capital":11700000.0,"capital":0.0,"interes":88952.05,"cuota":88952.05},
          {"op":7,"fecha":"2024-11-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":8,"fecha":"2024-12-05","saldo_capital":11700000.0,"capital":0.0,"interes":88952.05,"cuota":88952.05},
          {"op":9,"fecha":"2025-01-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":10,"fecha":"2025-02-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":11,"fecha":"2025-03-05","saldo_capital":11700000.0,"capital":0.0,"interes":83021.92,"cuota":83021.92},
          {"op":12,"fecha":"2025-04-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":13,"fecha":"2025-05-05","saldo_capital":11700000.0,"capital":0.0,"interes":88952.05,"cuota":88952.05},
          {"op":14,"fecha":"2025-06-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":15,"fecha":"2025-07-05","saldo_capital":11700000.0,"capital":0.0,"interes":88952.05,"cuota":88952.05},
          {"op":16,"fecha":"2025-08-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":17,"fecha":"2025-09-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":18,"fecha":"2025-10-05","saldo_capital":11700000.0,"capital":0.0,"interes":88952.05,"cuota":88952.05},
          {"op":19,"fecha":"2025-11-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":20,"fecha":"2025-12-05","saldo_capital":11700000.0,"capital":0.0,"interes":88952.05,"cuota":88952.05},
          {"op":21,"fecha":"2026-01-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":22,"fecha":"2026-02-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":23,"fecha":"2026-03-05","saldo_capital":11700000.0,"capital":0.0,"interes":83021.92,"cuota":83021.92},
          {"op":24,"fecha":"2026-04-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":25,"fecha":"2026-05-05","saldo_capital":11700000.0,"capital":0.0,"interes":88952.05,"cuota":88952.05},
          {"op":26,"fecha":"2026-06-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":27,"fecha":"2026-07-05","saldo_capital":11700000.0,"capital":0.0,"interes":88952.05,"cuota":88952.05},
          {"op":28,"fecha":"2026-08-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":29,"fecha":"2026-09-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":30,"fecha":"2026-10-05","saldo_capital":11700000.0,"capital":0.0,"interes":88952.05,"cuota":88952.05},
          {"op":31,"fecha":"2026-11-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":32,"fecha":"2026-12-05","saldo_capital":11700000.0,"capital":0.0,"interes":88952.05,"cuota":88952.05},
          {"op":33,"fecha":"2027-01-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":34,"fecha":"2027-02-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":35,"fecha":"2027-03-05","saldo_capital":11700000.0,"capital":0.0,"interes":83021.92,"cuota":83021.92},
          {"op":36,"fecha":"2027-04-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":37,"fecha":"2027-05-05","saldo_capital":11700000.0,"capital":0.0,"interes":88952.05,"cuota":88952.05},
          {"op":38,"fecha":"2027-06-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":39,"fecha":"2027-07-05","saldo_capital":11700000.0,"capital":0.0,"interes":88952.05,"cuota":88952.05},
          {"op":40,"fecha":"2027-08-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":41,"fecha":"2027-09-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":42,"fecha":"2027-10-05","saldo_capital":11700000.0,"capital":0.0,"interes":88952.05,"cuota":88952.05},
          {"op":43,"fecha":"2027-11-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":44,"fecha":"2027-12-05","saldo_capital":11700000.0,"capital":0.0,"interes":88952.05,"cuota":88952.05},
          {"op":45,"fecha":"2028-01-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":46,"fecha":"2028-02-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":47,"fecha":"2028-03-05","saldo_capital":11700000.0,"capital":0.0,"interes":85986.99,"cuota":85986.99},
          {"op":48,"fecha":"2028-04-05","saldo_capital":11700000.0,"capital":0.0,"interes":91917.12,"cuota":91917.12},
          {"op":49,"fecha":"2028-04-29","saldo_capital":0.0,"capital":11700000.0,"interes":71161.64,"cuota":11771161.64}
        ]
    }
}


# ─────────────────────────────────────────────
# LECTURA DINÁMICA DE PLAN TIERRA Y DIVIDENDOS DESDE EXCEL
# Fuente: PRESUPUESTO_PRESTAMOS_Y_TIERRA.xlsx
# Se lee en cada llamada al endpoint para reflejar cambios automáticamente.
# ─────────────────────────────────────────────

import glob as _glob

def _encontrar_excel_ppto():
    """Busca el archivo PRESUPUESTO en /data/sources/ con cualquier variante de nombre."""
    patrones = [
        "/data/sources/PRESUPUESTO, PRESTAMOS Y TIERRA.xlsx",
        "/data/sources/PRESUPUESTO_PRESTAMOS_Y_TIERRA.xlsx",
        "/data/sources/PRESUPUESTO PRESTAMOS Y TIERRA.xlsx",
        "/data/sources/PRESUPUESTO, PRESTAMOS*.xlsx",
        "/data/sources/PRESUPUESTO_PRESTAMOS*.xlsx",
        "/data/sources/PRESUPUESTO PRESTAMOS*.xlsx",
        "/data/sources/PRESUPUESTO*.xlsx",
    ]
    for p in patrones:
        matches = _glob.glob(p)
        if matches:
            return matches[0]
    return None

def _leer_plan_tierra_excel():
    try:
        from openpyxl import load_workbook
        ruta = _encontrar_excel_ppto()
        if not ruta:
            return []
        wb = load_workbook(ruta, read_only=True, data_only=True)
        ws = wb["PLAN DE PAGOS TIERRA"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        # Fila 0 = encabezados: col0=SOCIEDAD, col4=CONCEPTO, col6+=fechas
        header = rows[0]
        fecha_cols = []
        for i, h in enumerate(header):
            if i < 6: continue
            if hasattr(h, 'strftime'):
                fecha_cols.append((i, h.strftime("%Y-%m-%d")))
        resultado = []
        for row in rows[1:]:
            sociedad = str(row[0]).strip().upper() if row[0] else None
            if not sociedad or sociedad in ('NONE', 'TOTALES', 'SOCIEDAD'):
                continue
            concepto = str(row[4]).strip() if row[4] else ""
            if concepto.upper() not in ("TIERRA", "PAGO DE TIERRA", "UTILIDADES"):
                continue
            pagos = {}
            for col_i, fecha_str in fecha_cols:
                v = row[col_i] if col_i < len(row) else None
                if v and isinstance(v, (int, float)) and v > 0:
                    pagos[fecha_str] = float(v)
            if pagos:
                resultado.append({
                    "sociedad": sociedad,
                    "concepto": concepto,
                    "pagos": pagos,
                })
        wb.close()
        return resultado
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning(f"Error leyendo PLAN TIERRA desde Excel: {e}")
        return []

def _leer_plan_dividendos_excel():
    try:
        from openpyxl import load_workbook
        ruta = _encontrar_excel_ppto()
        if not ruta:
            return []
        wb = load_workbook(ruta, read_only=True, data_only=True)
        ws = wb["PLAN DE PAGOS UTILIDADES"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        # Fila 0 = encabezados: col0=SOCIEDAD, col2=CONCEPTO, col3=CTA SECCION, col4+=fechas
        header = rows[0]
        fecha_cols = []
        for i, h in enumerate(header):
            if i < 4: continue
            if hasattr(h, 'strftime'):
                fecha_cols.append((i, h.strftime("%Y-%m-%d")))
        resultado = []
        for row in rows[1:]:
            sociedad = str(row[0]).strip().upper() if row[0] else None
            if not sociedad or sociedad in ('NONE', 'TOTALES QTZ', 'SOCIEDAD'):
                continue
            concepto = str(row[2]).strip() if row[2] else ""
            cuenta   = str(row[3]).strip() if row[3] else concepto
            if not concepto or concepto.upper() in ('NONE', 'CONCEPTO'):
                continue
            pagos = {}
            for col_i, fecha_str in fecha_cols:
                v = row[col_i] if col_i < len(row) else None
                if v and isinstance(v, (int, float)) and v > 0:
                    pagos[fecha_str] = float(v)
            if pagos:
                resultado.append({
                    "sociedad": sociedad,
                    "cuenta": cuenta,
                    "pagos": pagos,
                })
        wb.close()
        return resultado
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning(f"Error leyendo PLAN DIVIDENDOS desde Excel: {e}")
        return []

def get_plan_tierra():
    """Retorna PLAN_TIERRA leyendo del Excel. Si falla, retorna lista vacía."""
    return _leer_plan_tierra_excel()

def get_plan_dividendos():
    """Retorna PLAN_DIVIDENDOS leyendo del Excel. Si falla, retorna lista vacía."""
    return _leer_plan_dividendos_excel()

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def calcular_tir(flujos, max_iter=200, tol=1e-6):
    """TIR con manejo de overflow."""
    if not flujos or all(f == 0 for f in flujos): return None
    if all(f >= 0 for f in flujos) or all(f <= 0 for f in flujos): return None
    try:
        tasa = 0.10
        for _ in range(max_iter):
            try:
                npv   = sum(f / (1+tasa)**i for i, f in enumerate(flujos))
                d_npv = sum(-i * f / (1+tasa)**(i+1) for i, f in enumerate(flujos))
            except (OverflowError, ZeroDivisionError):
                return None
            if d_npv == 0: break
            nueva = max(-0.99, min(tasa - npv / d_npv, 10.0))
            if abs(nueva - tasa) < tol: return round(nueva * 100, 2)
            tasa = nueva
        return round(tasa * 100, 2)
    except Exception:
        return None

def _calcular_tir_OLD(flujos, max_iter=200, tol=1e-6):
    if not flujos or all(f == 0 for f in flujos): return None
    if all(f >= 0 for f in flujos) or all(f <= 0 for f in flujos): return None
    tasa = 0.10
    for _ in range(max_iter):
        npv   = sum(f / (1+tasa)**i for i, f in enumerate(flujos))
        d_npv = sum(-i * f / (1+tasa)**(i+1) for i, f in enumerate(flujos))
        if d_npv == 0: break
        nueva = tasa - npv / d_npv
        if abs(nueva - tasa) < tol: return round(nueva * 100, 2)
        tasa = nueva
    return round(tasa * 100, 2)

def calcular_van(flujos, tasa):
    if tasa <= -1: return 0.0
    return sum(f / (1+tasa)**i for i, f in enumerate(flujos))

def get_ppto_key(empresa):
    """Obtiene la clave de PPTO para una empresa de la BD."""
    return EMPRESA_TO_PPTO.get(empresa, '')

def normalizar_sociedad_flujos(empresa_bd):
    """Convierte nombre empresa BD a formato que usa flujos_efectivo."""
    mapeo = {
        "Eficiencia Urbana, S. A.":     "EFICIENCIA URBANA",
        "Capipos, S.A.":                "CAPIPOS",
        "Ottavia, S.A.":                "OTTAVIA",
        "Vilet, S.A.":                  "VILET",
        "Tezzoli, S.A.":                "TEZZOLI",
        "Garbatella, S.A.":             "GARBATELLA",
        "Urviba 2, S.A.":               "URBIVA",
        "Utilica, S.A.":                "UTILICA",
        "Corcolle, S.A.":               "CORCOLLE",
        "Leofreni, S.A.":               "LEOFRENI",
        "Gibraleon, S.A.":              "GIBRALEON",
        "Talocci, S.A.":                "TALOCCI",
        "Ovest, S.A.":                  "OVEST",
        "Rossio, S.A.":                 "ROSSIO",
        "Servicios Generales CCC, S.A.":"SER GEN CCC",
        "Frugalex, S.A.":               "FRUGALEX",
    }
    return mapeo.get(empresa_bd, empresa_bd.upper())

def normalizar_empresa_cartera(empresa_bd):
    mapeo = {
        "Eficiencia Urbana, S. A.":      "Eficiencia Urbana",
        "Capipos, S.A.":                 "Capipos",
        "Ottavia, S.A.":                 "Ottavia",
        "Vilet, S.A.":                   "Vilet",
        "Tezzoli, S.A.":                 "Tezzoli",
        "Garbatella, S.A.":              "Garbatella",
        "Urviba 2, S.A.":                "Urbiva 2",
        "Utilica, S.A.":                 "Utilica",
        "Corcolle, S.A.":                "Corcolle",
        "Leofreni, S.A.":                "Leofreni",
        "Gibraleon, S.A.":               "Gibraleon",
        "Talocci, S.A.":                 "Talocci",
        "Ovest, S.A.":                   "Ovest",
        "Rossio, S.A.":                  "Rossio",
        "Servicios Generales CCC, S.A.": "Servicios Generales",
        "Frugalex, S.A.":                "Frugalex",
    }
    return mapeo.get(empresa_bd, empresa_bd)


# ─────────────────────────────────────────────
# ENDPOINTS
# ─────────────────────────────────────────────

@router.get("/")
async def get_proyectos(db: Session = Depends(get_db), user=Depends(get_current_user)):
    rows = db.execute(text("""
        SELECT DISTINCT p.nombre_sociedad, p.nombre_proyecto
        FROM proyectos p WHERE p.activo = true ORDER BY p.nombre_sociedad
    """)).fetchall()
    return [{"empresa": r.nombre_sociedad, "proyecto": r.nombre_proyecto} for r in rows]




# ═══════════ PRÉSTAMOS BANCARIOS — datos embebidos del Excel ═══════════
# Fuente: pestaña PRESTAMOS BANCARIOS del archivo PRESUPUESTO_PRESTAMOS_Y_TIERRA.xlsx
# 4 empresas: EU (60 cuotas/2030), Utilica (60/2030), Ser.Gen. (35/2027), Capipos (49/2028)
PRESTAMOS_SCHEDULE = [{"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 1, "fecha_pago": "2025-10-04", "saldo_capital": 14000000.0, "capital": 0.0, "interes": 57994.52, "cuota": 57994.52}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 2, "fecha_pago": "2025-11-04", "saldo_capital": 14000000.0, "capital": 0.0, "interes": 74909.59, "cuota": 74909.59}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 3, "fecha_pago": "2025-12-04", "saldo_capital": 14000000.0, "capital": 72413.79, "interes": 102526.03, "cuota": 174939.82}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 4, "fecha_pago": "2026-01-05", "saldo_capital": 13927586.21, "capital": 72413.79, "interes": 109154.66, "cuota": 181568.45}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 5, "fecha_pago": "2026-02-04", "saldo_capital": 13855172.42, "capital": 72413.79, "interes": 101916.39, "cuota": 174330.18}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 6, "fecha_pago": "2026-03-04", "saldo_capital": 13782758.63, "capital": 72413.79, "interes": 94123.33, "cuota": 166537.12}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 7, "fecha_pago": "2026-04-06", "saldo_capital": 13710344.84, "capital": 132413.79, "interes": 110887.48, "cuota": 243301.27}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 8, "fecha_pago": "2026-05-04", "saldo_capital": 13577931.05, "capital": 132413.79, "interes": 93743.52, "cuota": 226157.31}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 9, "fecha_pago": "2026-06-04", "saldo_capital": 13445517.26, "capital": 132413.79, "interes": 102775.32, "cuota": 235189.11}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 10, "fecha_pago": "2026-07-04", "saldo_capital": 13313103.47, "capital": 132413.79, "interes": 98480.49, "cuota": 230894.28}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 11, "fecha_pago": "2026-08-04", "saldo_capital": 13180689.68, "capital": 132413.79, "interes": 100751.03, "cuota": 233164.82}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 12, "fecha_pago": "2026-09-04", "saldo_capital": 13048275.89, "capital": 132413.79, "interes": 99738.88, "cuota": 232152.67}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 13, "fecha_pago": "2026-10-05", "saldo_capital": 12915862.1, "capital": 132413.79, "interes": 98726.73, "cuota": 231140.52}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 14, "fecha_pago": "2026-11-04", "saldo_capital": 12783448.31, "capital": 132413.79, "interes": 94562.49, "cuota": 226976.28}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 15, "fecha_pago": "2026-12-04", "saldo_capital": 12651034.52, "capital": 132413.79, "interes": 93583.0, "cuota": 225996.79}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 16, "fecha_pago": "2027-01-04", "saldo_capital": 12518620.73, "capital": 132413.79, "interes": 95690.28, "cuota": 228104.07}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 17, "fecha_pago": "2027-02-04", "saldo_capital": 12386206.94, "capital": 132413.79, "interes": 94678.13, "cuota": 227091.92}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 18, "fecha_pago": "2027-03-04", "saldo_capital": 12253793.15, "capital": 132413.79, "interes": 84601.53, "cuota": 217015.32}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 19, "fecha_pago": "2027-04-05", "saldo_capital": 12121379.36, "capital": 132413.79, "interes": 95642.66, "cuota": 228056.45}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 20, "fecha_pago": "2027-05-04", "saldo_capital": 11988965.57, "capital": 132413.79, "interes": 85729.32, "cuota": 218143.11}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 21, "fecha_pago": "2027-06-04", "saldo_capital": 11856551.78, "capital": 132413.79, "interes": 90629.53, "cuota": 223043.32}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 22, "fecha_pago": "2027-07-05", "saldo_capital": 11724137.99, "capital": 132413.79, "interes": 89617.38, "cuota": 222031.17}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 23, "fecha_pago": "2027-08-04", "saldo_capital": 11591724.2, "capital": 132413.79, "interes": 85747.0, "cuota": 218160.79}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 24, "fecha_pago": "2027-09-04", "saldo_capital": 11459310.41, "capital": 132413.79, "interes": 87593.09, "cuota": 220006.88}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 25, "fecha_pago": "2027-10-04", "saldo_capital": 11326896.62, "capital": 132413.79, "interes": 83788.0, "cuota": 216201.79}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 26, "fecha_pago": "2027-11-04", "saldo_capital": 11194482.83, "capital": 132413.79, "interes": 85568.79, "cuota": 217982.58}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 27, "fecha_pago": "2027-12-04", "saldo_capital": 11062069.04, "capital": 132413.79, "interes": 81829.0, "cuota": 214242.79}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 28, "fecha_pago": "2028-01-04", "saldo_capital": 10929655.25, "capital": 132413.79, "interes": 83544.49, "cuota": 215958.28}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 29, "fecha_pago": "2028-02-04", "saldo_capital": 10797241.46, "capital": 132413.79, "interes": 82532.34, "cuota": 214946.13}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 30, "fecha_pago": "2028-03-04", "saldo_capital": 10664827.67, "capital": 132413.79, "interes": 76260.82, "cuota": 208674.61}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 31, "fecha_pago": "2028-04-04", "saldo_capital": 10532413.88, "capital": 132413.79, "interes": 80508.04, "cuota": 212921.83}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 32, "fecha_pago": "2028-05-04", "saldo_capital": 10400000.09, "capital": 132413.79, "interes": 76931.51, "cuota": 209345.3}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 33, "fecha_pago": "2028-06-05", "saldo_capital": 10267586.3, "capital": 132413.79, "interes": 81015.48, "cuota": 213429.27}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 34, "fecha_pago": "2028-07-04", "saldo_capital": 10135172.51, "capital": 132413.79, "interes": 72473.43, "cuota": 204887.22}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 35, "fecha_pago": "2028-08-04", "saldo_capital": 10002758.72, "capital": 132413.79, "interes": 76459.44, "cuota": 208873.23}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 36, "fecha_pago": "2028-09-04", "saldo_capital": 9870344.93, "capital": 132413.79, "interes": 75447.29, "cuota": 207861.08}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 37, "fecha_pago": "2028-10-04", "saldo_capital": 9737931.14, "capital": 132413.79, "interes": 72034.01, "cuota": 204447.8}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 38, "fecha_pago": "2028-11-04", "saldo_capital": 9605517.35, "capital": 132413.79, "interes": 73423.0, "cuota": 205836.79}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 39, "fecha_pago": "2028-12-04", "saldo_capital": 9473103.56, "capital": 132413.79, "interes": 70075.01, "cuota": 202488.8}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 40, "fecha_pago": "2029-01-04", "saldo_capital": 9340689.77, "capital": 132413.79, "interes": 71398.7, "cuota": 203812.49}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 41, "fecha_pago": "2029-02-05", "saldo_capital": 9208275.98, "capital": 132413.79, "interes": 72657.08, "cuota": 205070.87}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 42, "fecha_pago": "2029-03-05", "saldo_capital": 9075862.19, "capital": 132413.79, "interes": 62660.75, "cuota": 195074.54}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 43, "fecha_pago": "2029-04-04", "saldo_capital": 8943448.4, "capital": 132413.79, "interes": 66157.02, "cuota": 198570.81}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 44, "fecha_pago": "2029-05-04", "saldo_capital": 8811034.61, "capital": 132413.79, "interes": 65177.52, "cuota": 197591.31}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 45, "fecha_pago": "2029-06-04", "saldo_capital": 8678620.82, "capital": 132413.79, "interes": 66337.95, "cuota": 198751.74}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 46, "fecha_pago": "2029-07-04", "saldo_capital": 8546207.03, "capital": 132413.79, "interes": 63218.52, "cuota": 195632.31}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 47, "fecha_pago": "2029-08-04", "saldo_capital": 8413793.24, "capital": 132413.79, "interes": 64313.65, "cuota": 196727.44}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 48, "fecha_pago": "2029-09-04", "saldo_capital": 8281379.45, "capital": 132413.79, "interes": 63301.5, "cuota": 195715.29}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 49, "fecha_pago": "2029-10-04", "saldo_capital": 8148965.66, "capital": 132413.79, "interes": 60280.02, "cuota": 192693.81}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 50, "fecha_pago": "2029-11-05", "saldo_capital": 8016551.87, "capital": 132413.79, "interes": 63253.89, "cuota": 195667.68}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 51, "fecha_pago": "2029-12-04", "saldo_capital": 7884138.08, "capital": 132413.79, "interes": 56376.99, "cuota": 188790.78}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 52, "fecha_pago": "2030-01-04", "saldo_capital": 7751724.29, "capital": 132413.79, "interes": 59252.91, "cuota": 191666.7}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 53, "fecha_pago": "2030-02-04", "saldo_capital": 7619310.5, "capital": 132413.79, "interes": 58240.76, "cuota": 190654.55}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 54, "fecha_pago": "2030-03-04", "saldo_capital": 7486896.71, "capital": 132413.79, "interes": 51690.36, "cuota": 184104.15}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 55, "fecha_pago": "2030-04-04", "saldo_capital": 7354482.92, "capital": 132413.79, "interes": 56216.46, "cuota": 188630.25}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 56, "fecha_pago": "2030-05-04", "saldo_capital": 7222069.13, "capital": 132413.79, "interes": 53423.53, "cuota": 185837.32}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 57, "fecha_pago": "2030-06-04", "saldo_capital": 7089655.34, "capital": 132413.79, "interes": 54192.16, "cuota": 186605.95}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 58, "fecha_pago": "2030-07-04", "saldo_capital": 6957241.55, "capital": 132413.79, "interes": 51464.53, "cuota": 183878.32}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 59, "fecha_pago": "2030-08-05", "saldo_capital": 6824827.76, "capital": 132413.79, "interes": 53850.7, "cuota": 186264.49}, {"sociedad": "EFICIENCIA URBANA", "sociedad_display": "Eficiencia Urbana", "no_operacion": 60, "fecha_pago": "2030-09-04", "saldo_capital": 6692413.97, "capital": 6692413.97, "interes": 49505.53, "cuota": 6741919.5}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 1, "fecha_pago": "2025-10-06", "saldo_capital": 5600000.0, "capital": 0.0, "interes": 53852.05, "cuota": 53852.05}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 2, "fecha_pago": "2025-11-05", "saldo_capital": 5600000.0, "capital": 0.0, "interes": 41424.66, "cuota": 41424.66}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 3, "fecha_pago": "2025-12-05", "saldo_capital": 5600000.0, "capital": 0.0, "interes": 41424.66, "cuota": 41424.66}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 4, "fecha_pago": "2026-01-05", "saldo_capital": 5600000.0, "capital": 0.0, "interes": 42805.48, "cuota": 42805.48}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 5, "fecha_pago": "2026-02-05", "saldo_capital": 5600000.0, "capital": 0.0, "interes": 42805.48, "cuota": 42805.48}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 6, "fecha_pago": "2026-03-05", "saldo_capital": 5600000.0, "capital": 0.0, "interes": 38663.01, "cuota": 38663.01}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 7, "fecha_pago": "2026-04-06", "saldo_capital": 5600000.0, "capital": 50000.0, "interes": 44186.3, "cuota": 94186.3}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 8, "fecha_pago": "2026-05-05", "saldo_capital": 5550000.0, "capital": 50000.0, "interes": 39686.27, "cuota": 89686.26999999999}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 9, "fecha_pago": "2026-06-05", "saldo_capital": 5500000.0, "capital": 50000.0, "interes": 42041.1, "cuota": 92041.1}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 10, "fecha_pago": "2026-07-06", "saldo_capital": 5450000.0, "capital": 50000.0, "interes": 41658.9, "cuota": 91658.9}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 11, "fecha_pago": "2026-08-05", "saldo_capital": 5400000.0, "capital": 50000.0, "interes": 39945.21, "cuota": 89945.20999999999}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 12, "fecha_pago": "2026-09-05", "saldo_capital": 5350000.0, "capital": 50000.0, "interes": 40894.52, "cuota": 90894.51999999999}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 13, "fecha_pago": "2026-10-05", "saldo_capital": 5300000.0, "capital": 75000.0, "interes": 39205.48, "cuota": 114205.48000000001}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 14, "fecha_pago": "2026-11-05", "saldo_capital": 5225000.0, "capital": 75000.0, "interes": 39939.04, "cuota": 114939.04000000001}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 15, "fecha_pago": "2026-12-05", "saldo_capital": 5150000.0, "capital": 75000.0, "interes": 38095.89, "cuota": 113095.89}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 16, "fecha_pago": "2027-01-05", "saldo_capital": 5075000.0, "capital": 75000.0, "interes": 38792.47, "cuota": 113792.47}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 17, "fecha_pago": "2027-02-05", "saldo_capital": 5000000.0, "capital": 75000.0, "interes": 38219.18, "cuota": 113219.18}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 18, "fecha_pago": "2027-03-05", "saldo_capital": 4925000.0, "capital": 75000.0, "interes": 34002.74, "cuota": 109002.73999999999}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 19, "fecha_pago": "2027-04-05", "saldo_capital": 4850000.0, "capital": 75000.0, "interes": 37072.6, "cuota": 112072.6}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 20, "fecha_pago": "2027-05-05", "saldo_capital": 4775000.0, "capital": 75000.0, "interes": 35321.92, "cuota": 110321.92}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 21, "fecha_pago": "2027-06-05", "saldo_capital": 4700000.0, "capital": 75000.0, "interes": 35926.03, "cuota": 110926.03}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 22, "fecha_pago": "2027-07-05", "saldo_capital": 4625000.0, "capital": 75000.0, "interes": 34212.33, "cuota": 109212.33}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 23, "fecha_pago": "2027-08-05", "saldo_capital": 4550000.0, "capital": 75000.0, "interes": 34779.45, "cuota": 109779.45}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 24, "fecha_pago": "2027-09-06", "saldo_capital": 4475000.0, "capital": 75000.0, "interes": 35309.59, "cuota": 110309.59}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 25, "fecha_pago": "2027-10-05", "saldo_capital": 4400000.0, "capital": 90000.0, "interes": 31463.01, "cuota": 121463.01}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 26, "fecha_pago": "2027-11-05", "saldo_capital": 4310000.0, "capital": 90000.0, "interes": 32944.93, "cuota": 122944.93}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 27, "fecha_pago": "2027-12-06", "saldo_capital": 4220000.0, "capital": 90000.0, "interes": 32256.99, "cuota": 122256.99}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 28, "fecha_pago": "2028-01-05", "saldo_capital": 4130000.0, "capital": 90000.0, "interes": 30550.68, "cuota": 120550.68}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 29, "fecha_pago": "2028-02-05", "saldo_capital": 4040000.0, "capital": 90000.0, "interes": 30881.1, "cuota": 120881.1}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 30, "fecha_pago": "2028-03-06", "saldo_capital": 3950000.0, "capital": 90000.0, "interes": 29219.18, "cuota": 119219.18}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 31, "fecha_pago": "2028-04-05", "saldo_capital": 3860000.0, "capital": 90000.0, "interes": 28553.42, "cuota": 118553.42}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 32, "fecha_pago": "2028-05-05", "saldo_capital": 3770000.0, "capital": 90000.0, "interes": 27887.67, "cuota": 117887.67}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 33, "fecha_pago": "2028-06-05", "saldo_capital": 3680000.0, "capital": 90000.0, "interes": 28129.32, "cuota": 118129.32}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 34, "fecha_pago": "2028-07-05", "saldo_capital": 3590000.0, "capital": 90000.0, "interes": 26556.16, "cuota": 116556.16}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 35, "fecha_pago": "2028-08-05", "saldo_capital": 3500000.0, "capital": 90000.0, "interes": 26753.42, "cuota": 116753.42}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 36, "fecha_pago": "2028-09-05", "saldo_capital": 3410000.0, "capital": 90000.0, "interes": 26065.48, "cuota": 116065.48}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 37, "fecha_pago": "2028-10-05", "saldo_capital": 3320000.0, "capital": 120000.0, "interes": 24558.9, "cuota": 144558.9}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 38, "fecha_pago": "2028-11-05", "saldo_capital": 3200000.0, "capital": 120000.0, "interes": 25249.32, "cuota": 145249.32}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 39, "fecha_pago": "2028-12-05", "saldo_capital": 3080000.0, "capital": 120000.0, "interes": 22024.11, "cuota": 142024.11}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 40, "fecha_pago": "2029-01-05", "saldo_capital": 2960000.0, "capital": 120000.0, "interes": 22625.75, "cuota": 142625.75}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 41, "fecha_pago": "2029-02-05", "saldo_capital": 2840000.0, "capital": 120000.0, "interes": 21708.49, "cuota": 141708.49}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 42, "fecha_pago": "2029-03-05", "saldo_capital": 2720000.0, "capital": 120000.0, "interes": 18779.18, "cuota": 138779.18}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 43, "fecha_pago": "2029-04-05", "saldo_capital": 2600000.0, "capital": 120000.0, "interes": 19873.97, "cuota": 139873.97}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 44, "fecha_pago": "2029-05-05", "saldo_capital": 2480000.0, "capital": 120000.0, "interes": 18345.21, "cuota": 138345.21}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 45, "fecha_pago": "2029-06-05", "saldo_capital": 2360000.0, "capital": 120000.0, "interes": 18039.45, "cuota": 138039.45}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 46, "fecha_pago": "2029-07-05", "saldo_capital": 2240000.0, "capital": 120000.0, "interes": 16569.86, "cuota": 136569.86}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 47, "fecha_pago": "2029-08-06", "saldo_capital": 2120000.0, "capital": 120000.0, "interes": 16727.67, "cuota": 136727.66999999998}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 48, "fecha_pago": "2029-09-05", "saldo_capital": 2000000.0, "capital": 120000.0, "interes": 14794.52, "cuota": 134794.52}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 49, "fecha_pago": "2029-10-05", "saldo_capital": 1880000.0, "capital": 156666.67, "interes": 13906.85, "cuota": 170573.52000000002}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 50, "fecha_pago": "2029-11-05", "saldo_capital": 1723333.33, "capital": 156666.67, "interes": 13172.88, "cuota": 169839.55000000002}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 51, "fecha_pago": "2029-12-05", "saldo_capital": 1566666.66, "capital": 156666.67, "interes": 11589.04, "cuota": 168255.71000000002}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 52, "fecha_pago": "2030-01-05", "saldo_capital": 1409999.99, "capital": 156666.67, "interes": 10777.81, "cuota": 167444.48}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 53, "fecha_pago": "2030-02-05", "saldo_capital": 1253333.32, "capital": 156666.67, "interes": 9580.27, "cuota": 166246.94}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 54, "fecha_pago": "2030-03-05", "saldo_capital": 1096666.65, "capital": 156666.67, "interes": 7571.51, "cuota": 164238.18000000002}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 55, "fecha_pago": "2030-04-05", "saldo_capital": 939999.98, "capital": 156666.67, "interes": 7185.21, "cuota": 163851.88}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 56, "fecha_pago": "2030-05-06", "saldo_capital": 783333.31, "capital": 156666.67, "interes": 5987.67, "cuota": 162654.34000000003}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 57, "fecha_pago": "2030-06-05", "saldo_capital": 626666.64, "capital": 156666.67, "interes": 4635.62, "cuota": 161302.29}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 58, "fecha_pago": "2030-07-05", "saldo_capital": 469999.97, "capital": 156666.67, "interes": 3476.71, "cuota": 160143.38}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 59, "fecha_pago": "2030-08-05", "saldo_capital": 313333.3, "capital": 156666.67, "interes": 2395.07, "cuota": 159061.74000000002}, {"sociedad": "UTILICA", "sociedad_display": "Utilica", "no_operacion": 60, "fecha_pago": "2030-09-05", "saldo_capital": 156666.63, "capital": 156666.63, "interes": 1197.53, "cuota": 157864.16}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 2, "fecha_pago": "2024-05-06", "saldo_capital": 3000000.0, "capital": 26349.24, "interes": 11761.64, "cuota": 38110.880000000005}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 3, "fecha_pago": "2024-06-05", "saldo_capital": 2973650.76, "capital": 26349.24, "interes": 11196.87, "cuota": 37546.11}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 4, "fecha_pago": "2024-07-05", "saldo_capital": 2947301.52, "capital": 26349.24, "interes": 11001.96, "cuota": 37351.2}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 5, "fecha_pago": "2024-08-05", "saldo_capital": 2920952.28, "capital": 25000.0, "interes": 21094.4, "cuota": 46094.4}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 6, "fecha_pago": "2024-09-05", "saldo_capital": 2895952.28, "capital": 25000.0, "interes": 20903.31, "cuota": 45903.31}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 7, "fecha_pago": "2024-10-05", "saldo_capital": 2870952.28, "capital": 25000.0, "interes": 21533.07, "cuota": 46533.07}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 8, "fecha_pago": "2024-11-05", "saldo_capital": 2845952.28, "capital": 25000.0, "interes": 22049.88, "cuota": 47049.880000000005}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 9, "fecha_pago": "2024-12-05", "saldo_capital": 2820952.28, "capital": 25000.0, "interes": 21163.21, "cuota": 46163.21}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 10, "fecha_pago": "2025-01-06", "saldo_capital": 2795952.28, "capital": 25000.0, "interes": 22357.1, "cuota": 47357.1}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 11, "fecha_pago": "2025-02-05", "saldo_capital": 2770952.28, "capital": 106731.0, "interes": 20793.35, "cuota": 127524.35}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 12, "fecha_pago": "2025-03-05", "saldo_capital": 2664221.28, "capital": 106731.0, "interes": 18689.97, "cuota": 125420.97}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 13, "fecha_pago": "2025-04-05", "saldo_capital": 2557490.28, "capital": 106731.0, "interes": 19844.93, "cuota": 126575.93}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 14, "fecha_pago": "2025-05-05", "saldo_capital": 2450759.28, "capital": 106731.0, "interes": 18424.79, "cuota": 125155.79000000001}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 15, "fecha_pago": "2025-06-05", "saldo_capital": 2344028.28, "capital": 106731.0, "interes": 18213.26, "cuota": 124944.26}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 16, "fecha_pago": "2025-07-05", "saldo_capital": 2237297.28, "capital": 106731.0, "interes": 16845.76, "cuota": 123576.76}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 17, "fecha_pago": "2025-08-05", "saldo_capital": 2130566.28, "capital": 106731.0, "interes": 16581.59, "cuota": 123312.59}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 18, "fecha_pago": "2025-09-05", "saldo_capital": 2023835.28, "capital": 106731.0, "interes": 15765.75, "cuota": 122496.75}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 19, "fecha_pago": "2025-10-06", "saldo_capital": 1917104.28, "capital": 106731.0, "interes": 14949.92, "cuota": 121680.92}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 20, "fecha_pago": "2025-11-05", "saldo_capital": 1810373.28, "capital": 106731.0, "interes": 13687.69, "cuota": 120418.69}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 21, "fecha_pago": "2025-12-05", "saldo_capital": 1703642.28, "capital": 106731.0, "interes": 12898.18, "cuota": 119629.18}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 22, "fecha_pago": "2026-01-05", "saldo_capital": 1596911.28, "capital": 106731.0, "interes": 12502.42, "cuota": 119233.42}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 23, "fecha_pago": "2026-02-05", "saldo_capital": 1490180.28, "capital": 106731.0, "interes": 11686.58, "cuota": 118417.58}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 24, "fecha_pago": "2026-03-05", "saldo_capital": 1383449.28, "capital": 106731.0, "interes": 9847.38, "cuota": 116578.38}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 25, "fecha_pago": "2026-04-06", "saldo_capital": 1276718.28, "capital": 106731.0, "interes": 10369.72, "cuota": 117100.72}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 26, "fecha_pago": "2026-05-05", "saldo_capital": 1169987.28, "capital": 106731.0, "interes": 8662.1, "cuota": 115393.1}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 27, "fecha_pago": "2026-06-05", "saldo_capital": 1063256.28, "capital": 106731.0, "interes": 8423.25, "cuota": 115154.25}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 28, "fecha_pago": "2026-07-06", "saldo_capital": 956525.28, "capital": 106731.0, "interes": 7607.41, "cuota": 114338.41}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 29, "fecha_pago": "2026-08-05", "saldo_capital": 849794.28, "capital": 106731.0, "interes": 6582.04, "cuota": 113313.04}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 30, "fecha_pago": "2026-09-05", "saldo_capital": 743063.28, "capital": 106731.0, "interes": 5975.74, "cuota": 112706.74}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 31, "fecha_pago": "2026-10-05", "saldo_capital": 636332.28, "capital": 106731.0, "interes": 5003.01, "cuota": 111734.01}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 32, "fecha_pago": "2026-11-05", "saldo_capital": 529601.28, "capital": 106731.0, "interes": 4344.08, "cuota": 111075.08}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 33, "fecha_pago": "2026-12-05", "saldo_capital": 422870.28, "capital": 106731.0, "interes": 3423.97, "cuota": 110154.97}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 34, "fecha_pago": "2027-01-05", "saldo_capital": 316139.28, "capital": 106731.0, "interes": 2712.41, "cuota": 109443.41}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 35, "fecha_pago": "2027-02-05", "saldo_capital": 209408.28, "capital": 105152.41, "interes": 1896.57, "cuota": 107048.98000000001}, {"sociedad": "SER GEN CCC", "sociedad_display": "Servicios Generales", "no_operacion": 36, "fecha_pago": "2027-03-05", "saldo_capital": 104255.87, "capital": 104255.87, "interes": 1015.69, "cuota": 105271.56}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 1, "fecha_pago": "2024-05-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 14825.34, "cuota": 14825.34}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 2, "fecha_pago": "2024-06-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 3, "fecha_pago": "2024-07-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 88952.05, "cuota": 88952.05}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 4, "fecha_pago": "2024-08-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 5, "fecha_pago": "2024-09-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 6, "fecha_pago": "2024-10-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 88952.05, "cuota": 88952.05}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 7, "fecha_pago": "2024-11-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 8, "fecha_pago": "2024-12-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 88952.05, "cuota": 88952.05}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 9, "fecha_pago": "2025-01-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 10, "fecha_pago": "2025-02-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 11, "fecha_pago": "2025-03-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 83021.92, "cuota": 83021.92}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 12, "fecha_pago": "2025-04-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 13, "fecha_pago": "2025-05-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 88952.05, "cuota": 88952.05}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 14, "fecha_pago": "2025-06-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 15, "fecha_pago": "2025-07-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 88952.05, "cuota": 88952.05}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 16, "fecha_pago": "2025-08-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 17, "fecha_pago": "2025-09-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 18, "fecha_pago": "2025-10-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 88952.05, "cuota": 88952.05}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 19, "fecha_pago": "2025-11-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 20, "fecha_pago": "2025-12-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 88952.05, "cuota": 88952.05}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 21, "fecha_pago": "2026-01-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 22, "fecha_pago": "2026-02-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 23, "fecha_pago": "2026-03-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 83021.92, "cuota": 83021.92}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 24, "fecha_pago": "2026-04-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 25, "fecha_pago": "2026-05-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 88952.05, "cuota": 88952.05}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 26, "fecha_pago": "2026-06-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 27, "fecha_pago": "2026-07-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 88952.05, "cuota": 88952.05}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 28, "fecha_pago": "2026-08-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 29, "fecha_pago": "2026-09-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 30, "fecha_pago": "2026-10-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 88952.05, "cuota": 88952.05}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 31, "fecha_pago": "2026-11-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 32, "fecha_pago": "2026-12-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 88952.05, "cuota": 88952.05}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 33, "fecha_pago": "2027-01-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 34, "fecha_pago": "2027-02-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 35, "fecha_pago": "2027-03-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 83021.92, "cuota": 83021.92}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 36, "fecha_pago": "2027-04-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 37, "fecha_pago": "2027-05-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 88952.05, "cuota": 88952.05}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 38, "fecha_pago": "2027-06-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 39, "fecha_pago": "2027-07-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 88952.05, "cuota": 88952.05}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 40, "fecha_pago": "2027-08-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 41, "fecha_pago": "2027-09-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 42, "fecha_pago": "2027-10-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 88952.05, "cuota": 88952.05}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 43, "fecha_pago": "2027-11-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 44, "fecha_pago": "2027-12-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 88952.05, "cuota": 88952.05}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 45, "fecha_pago": "2028-01-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 46, "fecha_pago": "2028-02-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 47, "fecha_pago": "2028-03-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 85986.99, "cuota": 85986.99}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 48, "fecha_pago": "2028-04-05", "saldo_capital": 11700000.0, "capital": 0.0, "interes": 91917.12, "cuota": 91917.12}, {"sociedad": "CAPIPOS", "sociedad_display": "Capipos", "no_operacion": 49, "fecha_pago": "2028-04-29", "saldo_capital": 0.0, "capital": 11700000.0, "interes": 71161.64, "cuota": 11771161.64}]

def _get_prestamos_empresa(sociedad_flujos: str):
    """Filtra cuotas de préstamo para la empresa dada (partial match)."""
    soc_up = sociedad_flujos.upper()
    return [
        p for p in PRESTAMOS_SCHEDULE
        if soc_up in p['sociedad'] or p['sociedad'] in soc_up
    ]

@router.get("/{empresa}/prestamos")
async def get_prestamos_empresa(
    empresa: str,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """
    Devuelve el estado del préstamo bancario para la empresa:
    - pagado: cuotas con fecha_pago <= hoy (ejecutado real)
    - pendiente: cuotas con fecha_pago > hoy (proyección)
    - resumen: totales de capital, interés, cuota pagada vs pendiente
    """
    from datetime import date
    hoy = date.today().isoformat()
    sociedad_flujos = normalizar_sociedad_flujos(empresa)
    cuotas = _get_prestamos_empresa(sociedad_flujos)

    if not cuotas:
        return {
            "tiene_prestamo": False,
            "empresa": empresa,
            "cuotas": [],
            "resumen": {}
        }

    pagadas    = [c for c in cuotas if c['fecha_pago'] <= hoy]
    pendientes = [c for c in cuotas if c['fecha_pago'] >  hoy]

    resumen = {
        "monto_original":       cuotas[0]['saldo_capital'] if cuotas else 0,
        "saldo_actual":         cuotas[-1]['saldo_capital'] if pendientes else 0,
        "total_cuotas":         len(cuotas),
        "cuotas_pagadas":       len(pagadas),
        "cuotas_pendientes":    len(pendientes),
        # Ejecutado (pagado a hoy)
        "capital_pagado":       round(sum(c['capital'] for c in pagadas), 2),
        "intereses_pagados":    round(sum(c['interes']  for c in pagadas), 2),
        "cuota_pagada_total":   round(sum(c['cuota']    for c in pagadas), 2),
        # Pendiente (proyección)
        "capital_pendiente":    round(sum(c['capital'] for c in pendientes), 2),
        "intereses_pendientes": round(sum(c['interes']  for c in pendientes), 2),
        "cuota_pendiente_total":round(sum(c['cuota']    for c in pendientes), 2),
        # Próxima cuota
        "proxima_fecha":        pendientes[0]['fecha_pago']  if pendientes else None,
        "proxima_cuota":        round(pendientes[0]['cuota'] if pendientes else 0, 2),
        "proxima_capital":      round(pendientes[0]['capital'] if pendientes else 0, 2),
        "proximos_intereses":   round(pendientes[0]['interes'] if pendientes else 0, 2),
        # Vencimiento
        "fecha_vencimiento":    cuotas[-1]['fecha_pago'] if cuotas else None,
    }

    return {
        "tiene_prestamo": True,
        "empresa": empresa,
        "sociedad": sociedad_flujos,
        "resumen": resumen,
        "pagadas":    pagadas[-6:],   # últimas 6 cuotas pagadas
        "pendientes": pendientes[:12], # próximas 12 cuotas
    }

@router.get("/{empresa}")
async def get_supuestos(empresa: str, db: Session = Depends(get_db), user=Depends(get_current_user)):
    row = db.execute(text(
        "SELECT * FROM proyeccion_supuestos WHERE empresa = :e LIMIT 1"
    ), {"e": empresa}).fetchone()
    if not row:
        return {"empresa": empresa, "existe": False}
    return {
        "empresa": empresa, "existe": True,
        "ticket_proyectado": float(row.ticket_proyectado or 0),
        "tasa_interes":      float(row.tasa_interes or 0),
        "plazo_meses":       int(row.plazo_meses or 0),
        "anos_proyecto":     int(row.anos_proyecto or 5),
        "tasa_descuento":    float(row.tasa_descuento or 0.12),
        "pct_isr":           float(row.pct_isr or 0),
        "plazos_venta":      row.egresos_operativos or [],  # reutilizamos campo para plazos
        "anos_ic":           getattr(row, 'anos_ic', 0) or 0,
        "anos_egr":          getattr(row, 'anos_egr', 0) or 0,
        "actualizado_por":   row.actualizado_por,
        "actualizado_en":    str(row.actualizado_en) if row.actualizado_en else None,
    }


@router.post("/{empresa}")
async def save_supuestos(empresa: str, body: dict, db: Session = Depends(get_db), user=Depends(get_current_user)):
    db.execute(text("""
        INSERT INTO proyeccion_supuestos
          (empresa, ticket_proyectado, tasa_interes, plazo_meses,
           anos_proyecto, tasa_descuento, pct_isr,
           egresos_operativos, prestamos, pagos_tierra,
           anos_ic, anos_egr,
           actualizado_por, actualizado_en)
        VALUES (:e, :tk, :ti, :pm, :ap, :td, :isr,
                CAST(:eo AS jsonb), CAST(:pr AS jsonb), CAST(:pt AS jsonb),
                :anos_ic_val, :anos_egr_val, :who, NOW())
        ON CONFLICT (empresa) DO UPDATE SET
          ticket_proyectado  = EXCLUDED.ticket_proyectado,
          tasa_interes       = EXCLUDED.tasa_interes,
          plazo_meses        = EXCLUDED.plazo_meses,
          anos_proyecto      = EXCLUDED.anos_proyecto,
          tasa_descuento     = EXCLUDED.tasa_descuento,
          pct_isr            = EXCLUDED.pct_isr,
          egresos_operativos = EXCLUDED.egresos_operativos,
          prestamos          = EXCLUDED.prestamos,
          pagos_tierra       = EXCLUDED.pagos_tierra,
          anos_ic            = EXCLUDED.anos_ic,
          anos_egr           = EXCLUDED.anos_egr,
          actualizado_por    = EXCLUDED.actualizado_por,
          actualizado_en     = NOW()
    """), {
        "e":   empresa,
        "tk":  body.get("ticket_proyectado"),
        "ti":  body.get("tasa_interes"),
        "pm":  body.get("plazo_meses"),
        "ap":  body.get("anos_proyecto", 5),
        "td":  body.get("tasa_descuento", 0.12),
        "isr": body.get("pct_isr", 0),
        "eo":  json.dumps(body.get("plazos_venta", [])),
        "anos_ic_val": body.get("anos_ic", 0),
        "anos_egr_val": body.get("anos_egr", 0),
        "pr":  json.dumps(body.get("prestamos_manual", [])),
        "pt":  json.dumps(body.get("pagos_tierra_manual", [])),
        "who": getattr(user, "email", getattr(user, "nombre", "sistema")),
    })
    db.commit()
    return {"ok": True}


@router.get("/{empresa}/horizonte")
async def get_horizonte(empresa: str, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Calcula el horizonte automático basado en la última fecha de cobro en cartera."""
    row = db.execute(text("""
        SELECT
            MAX(EXTRACT(YEAR FROM fecha_programada_cobro)) AS ultimo_anio,
            MIN(EXTRACT(YEAR FROM fecha_programada_cobro)) AS primer_anio
        FROM ov_cartera
        WHERE empresa = :e
          AND line_status = 'O'
          AND fecha_programada_cobro IS NOT NULL
    """), {"e": normalizar_empresa_cartera(empresa)}).fetchone()

    anio_actual = date.today().year
    ultimo_anio = int(row.ultimo_anio or anio_actual)
    horizonte_calc = max(1, ultimo_anio - anio_actual + 1)

    return {
        "empresa": empresa,
        "ultimo_anio_ov": ultimo_anio,
        "anio_actual": anio_actual,
        "horizonte_calculado": horizonte_calc,
    }


@router.get("/{empresa}/ingresos-por-anio")
async def get_ingresos_por_anio(empresa: str, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Distribución anual de ingresos reales por fecha_programada_cobro."""
    rows = db.execute(text("""
        SELECT
            EXTRACT(YEAR FROM fecha_programada_cobro) AS anio,
            SUM(saldo_pendiente)                       AS total
        FROM ov_cartera
        WHERE empresa = :e
          AND line_status = 'O'
          AND fecha_programada_cobro IS NOT NULL
        GROUP BY anio
        ORDER BY anio
    """), {"e": normalizar_empresa_cartera(empresa)}).fetchall()

    return {
        "empresa": empresa,
        "distribucion": [{"anio": int(r.anio), "total": float(r.total)} for r in rows],
    }


@router.get("/{empresa}/ingresos")
async def get_ingresos(
    empresa: str,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Ingresos reales = saldo pendiente cartera + ingresos proyectados por plazos de venta."""

    # 1. Saldo pendiente real de cartera (ov_cartera)
    rr = db.execute(text("""
        SELECT
            COUNT(DISTINCT doc_num)                         AS contratos,
            COALESCE(SUM(saldo_pendiente), 0)               AS saldo_pendiente_total,
            COALESCE(SUM(CASE WHEN tipo_linea = 'BB' THEN saldo_pendiente ELSE 0 END), 0) AS saldo_capital,
            COALESCE(SUM(CASE WHEN tipo_linea = 'S'  THEN saldo_pendiente ELSE 0 END), 0) AS saldo_interes
        FROM ov_cartera
        WHERE empresa = :e
          AND line_status = 'O'
    """), {"e": normalizar_empresa_cartera(empresa)}).fetchone()

    ingresos_reales = {
        "contratos":        int(rr.contratos or 0),
        "saldo_pendiente":  float(rr.saldo_pendiente_total or 0),
        "saldo_capital":    float(rr.saldo_capital or 0),
        "saldo_interes":    float(rr.saldo_interes or 0),
    }

    # 2. Lotes disponibles (para tabla de plazos proyectados)
    lotes = db.execute(text("""
        SELECT l.unidad_key, p.nombre_proyecto,
               COALESCE(l.precio_con_descuento, l.precio_sin_descuento, 0) AS precio
        FROM lotes l
        JOIN proyectos p ON p.id = l.proyecto_id
        WHERE p.nombre_sociedad = :e AND l.estatus = 'DISPONIBLE'
        ORDER BY l.unidad_key
    """), {"e": empresa}).fetchall()

    disponibles = [{"lote": r.unidad_key, "proyecto": r.nombre_proyecto,
                    "precio": float(r.precio)} for r in lotes]

    return {
        "empresa": empresa,
        "ingresos_reales": ingresos_reales,
        "disponibles": disponibles,
        "total_disponibles": len(disponibles),
        "precio_promedio": round(sum(d["precio"] for d in disponibles) / len(disponibles), 2) if disponibles else 0,
    }


@router.get("/{empresa}/egresos-operativos")
async def get_egresos_operativos(
    empresa: str,
    anos_egr: int = Query(0, description="Años para distribuir egresos pendientes (0=no distribuir)"),
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Presupuesto vs Ejecutado para Urbanización y Administración."""
    ppto_key = get_ppto_key(empresa)
    ppto = PPTO_PROYECTOS.get(ppto_key, {"urb": 0, "adm": 0, "total": 0})
    sociedad_flujos = normalizar_sociedad_flujos(empresa)

    # Ejecutado base desde flujos_efectivo
    rows = db.execute(text("""
        SELECT
          CASE
            WHEN seccion IN ('EGRESOS / URBANIZACION', 'EGRESOS / MOVIMIENTO DE TIERRAS')
              THEN 'EGRESOS / URBANIZACION'
            ELSE seccion
          END AS seccion,
          COALESCE(SUM(monto_egreso), 0) AS ejecutado
        FROM flujos_efectivo
        WHERE sociedad ILIKE :s
          AND seccion IN ('EGRESOS / URBANIZACION', 'EGRESOS / MOVIMIENTO DE TIERRAS', 'EGRESOS / ADMINISTRACION')
        GROUP BY 1
    """), {"s": f"%{sociedad_flujos}%"}).fetchall()

    ejecutado = {"EGRESOS / URBANIZACION": 0.0, "EGRESOS / ADMINISTRACION": 0.0}
    for r in rows:
        ejecutado[r.seccion] = float(r.ejecutado)

    # Aplicar reclasificaciones (misma lógica que flujos.py)
    reclas = db.execute(text("""
        SELECT seccion_origen, seccion_destino, SUM(monto) AS monto
        FROM flujos_reclasificaciones
        WHERE sociedad ILIKE :s
          AND (seccion_origen IN ('EGRESOS / URBANIZACION','EGRESOS / MOVIMIENTO DE TIERRAS','EGRESOS / ADMINISTRACION')
           OR  seccion_destino IN ('EGRESOS / URBANIZACION','EGRESOS / MOVIMIENTO DE TIERRAS','EGRESOS / ADMINISTRACION'))
        GROUP BY seccion_origen, seccion_destino
    """), {"s": f"%{sociedad_flujos}%"}).fetchall()

    for r in reclas:
        ori_raw, dst_raw, monto = r.seccion_origen, r.seccion_destino, float(r.monto or 0)
        if not monto:
            continue
        # Treat MOVIMIENTO DE TIERRAS as URBANIZACION for projection purposes
        ori = 'EGRESOS / URBANIZACION' if ori_raw == 'EGRESOS / MOVIMIENTO DE TIERRAS' else ori_raw
        dst = 'EGRESOS / URBANIZACION' if dst_raw == 'EGRESOS / MOVIMIENTO DE TIERRAS' else dst_raw
        if ori == dst:
            continue
        if ori in ejecutado:
            ejecutado[ori] -= monto
        if dst in ejecutado:
            ejecutado[dst] += monto

    ejec_urb   = max(ejecutado["EGRESOS / URBANIZACION"],   0.0)
    ejec_adm   = max(ejecutado["EGRESOS / ADMINISTRACION"], 0.0)
    ejec_total = ejec_urb + ejec_adm

    ppto_urb = ppto["urb"]
    ppto_adm = ppto["adm"]
    ppto_total = ppto["total"]

    pendiente_urb = round(max(ppto_urb - ejec_urb, 0), 2)
    pendiente_adm = round(max(ppto_adm - ejec_adm, 0), 2)
    pendiente_total = round(max(ppto_total - ejec_total, 0), 2)

    # Distribución anual del pendiente si se solicitó
    dist_urb = round(pendiente_urb / anos_egr, 2) if anos_egr > 0 else None
    dist_adm = round(pendiente_adm / anos_egr, 2) if anos_egr > 0 else None
    dist_total = round(pendiente_total / anos_egr, 2) if anos_egr > 0 else None

    return {
        "empresa": empresa,
        "ppto_key": ppto_key,
        "anos_distribucion": anos_egr,
        "urbanizacion": {
            "presupuesto": round(ppto_urb, 2),
            "ejecutado":   round(ejec_urb, 2),
            "pendiente":   pendiente_urb,
            "avance_pct":  round(ejec_urb / ppto_urb * 100, 1) if ppto_urb else 0,
            "dist_anual":  dist_urb,
        },
        "administracion": {
            "presupuesto": round(ppto_adm, 2),
            "ejecutado":   round(ejec_adm, 2),
            "pendiente":   pendiente_adm,
            "avance_pct":  round(ejec_adm / ppto_adm * 100, 1) if ppto_adm else 0,
            "dist_anual":  dist_adm,
        },
        "total": {
            "presupuesto": round(ppto_total, 2),
            "ejecutado":   round(ejec_total, 2),
            "pendiente":   pendiente_total,
            "avance_pct":  round(ejec_total / ppto_total * 100, 1) if ppto_total else 0,
            "dist_anual":  dist_total,
        },
    }


@router.get("/{empresa}/egresos-financieros")
async def get_egresos_financieros(
    empresa: str,
    anos_ic: int = Query(0, description="Años para distribuir intercompany (0=no distribuir)"),
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Préstamos bancarios (tabla amortización vs pagado) e Intercompany."""
    sociedad_flujos = normalizar_sociedad_flujos(empresa)
    hoy = date.today()

    # ── PRÉSTAMO BANCARIO ──
    prestamo_info = PRESTAMOS_AMORT.get(empresa)
    prestamo_result = None

    if prestamo_info:
        # Pagado real desde flujos_efectivo (Prestamo Bancario)
        # Pagado real = Prestamo Bancario base (flujos_efectivo)
        # + Liberaciones y Cuota Capital desde reclasificaciones (subseccion level)
        # + Intereses Prestamo desde reclasificaciones (reclasificados desde URBA)
        pagado_row = db.execute(text("""
            SELECT COALESCE(SUM(monto_egreso), 0) AS pagado
            FROM flujos_efectivo
            WHERE sociedad ILIKE :s
              AND seccion = 'FINANCIAMIENTO'
              AND nombre_categoria = 'Prestamo Bancario'
        """), {"s": f"%{sociedad_flujos}%"}).fetchone()
        pagado_base = float(pagado_row.pagado or 0)

        # Add reclasificaciones detail: Cuota Capital + Liberaciones + Intereses Prestamo
        # En BD: subseccion='Prestamo Bancario', nombre_destino='Cuota Capital'/'Liberaciones'/'Intereses Prestamo'
        reclas_detail = db.execute(text("""
            SELECT nombre_destino AS tipo, SUM(monto) AS total
            FROM flujos_reclasificaciones
            WHERE sociedad ILIKE :s
              AND seccion_destino = 'FINANCIAMIENTO'
              AND subseccion = 'Prestamo Bancario'
              AND nombre_destino IN ('Cuota Capital', 'Liberaciones', 'Intereses Prestamo')
            GROUP BY nombre_destino
        """), {"s": f"%{sociedad_flujos}%"}).fetchall()

        reclas_por_subsec = {r.tipo: float(r.total or 0) for r in reclas_detail}
        capital_reclas    = reclas_por_subsec.get('Cuota Capital', 0) + reclas_por_subsec.get('Liberaciones', 0)
        intereses_reclas  = reclas_por_subsec.get('Intereses Prestamo', 0)

        # pagado_total para cruce con tabla amortizacion:
        # Cuota = Capital (Cuota Capital + Liberaciones) + Intereses
        # Capital pagado segun tabla = lo que flujo_efectivo registra como Prestamo Bancario
        # mas las liberaciones que son desembolsos/pagos extra no en la cuota regular
        pagado_total = pagado_base + capital_reclas + intereses_reclas

        cuotas = prestamo_info.get("cuotas", [])
        monto_original = prestamo_info["monto_original"]

        # Cuotas pagadas = todas con fecha <= hoy
        capital_pagado_tabla = 0.0
        interes_pagado_tabla = 0.0
        cuotas_pagadas = []
        cuotas_pendientes_raw = []
        for c in cuotas:
            try:
                fecha_cuota = datetime.strptime(c["fecha"], "%Y-%m-%d").date()
            except Exception:
                fecha_cuota = hoy + timedelta(days=1)
            if fecha_cuota <= hoy:
                capital_pagado_tabla += c["capital"]
                interes_pagado_tabla += c["interes"]
                cuotas_pagadas.append(c)
            else:
                cuotas_pendientes_raw.append(c)

        # Separar vencidas (fecha <= hoy) de futuras y acumular vencidas en un solo pago
        cuotas_vencidas = [c for c in cuotas_pendientes_raw
                           if datetime.strptime(c["fecha"], "%Y-%m-%d").date() <= hoy]
        cuotas_futuras  = [c for c in cuotas_pendientes_raw
                           if datetime.strptime(c["fecha"], "%Y-%m-%d").date() > hoy]

        cuotas_pendientes = []
        if cuotas_vencidas:
            cap_venc = sum(c["capital"] for c in cuotas_vencidas)
            int_venc = sum(c["interes"]  for c in cuotas_vencidas)
            prox_fecha = cuotas_futuras[0]["fecha"] if cuotas_futuras else hoy.replace(day=5).strftime("%Y-%m-%d")
            cuotas_pendientes.append({
                "op": "VENCIDO",
                "fecha": prox_fecha,
                "saldo_capital": cuotas_vencidas[0]["saldo_capital"] if cuotas_vencidas else 0,
                "capital": round(cap_venc, 2),
                "interes": round(int_venc, 2),
                "cuota": round(cap_venc + int_venc, 2),
                "es_vencido": True,
                "cuotas_acumuladas": len(cuotas_vencidas),
            })
        cuotas_pendientes.extend(cuotas_futuras)

        pendiente_capital = sum(c["capital"] for c in cuotas_pendientes)
        pendiente_interes = sum(c["interes"] for c in cuotas_pendientes)

        prestamo_result = {
            "banco":             prestamo_info["banco"],
            "no_credito":        prestamo_info["no_credito"],
            "monto_original":    monto_original,
            "tasa":              prestamo_info["tasa"],
            "pagado_capital":    round(capital_pagado_tabla, 2),
            "pagado_interes":    round(interes_pagado_tabla, 2),
            "pagado_total":      round(pagado_total, 2),
            "pendiente_capital": round(pendiente_capital, 2),
            "pendiente_interes": round(pendiente_interes, 2),
            "pendiente_total":   round(pendiente_capital + pendiente_interes, 2),
            "cuotas_pagadas":    len(cuotas_pagadas),
            "cuotas_pendientes": cuotas_pendientes,
            # Detalle de lo pagado por subsección (desde reclasificaciones)
            "pagado_flujo_base":     round(pagado_base, 2),
            "pagado_cuota_capital":  round(reclas_por_subsec.get('Cuota Capital', 0), 2),
            "pagado_liberaciones":   round(reclas_por_subsec.get('Liberaciones', 0), 2),
            "pagado_intereses":      round(intereses_reclas, 2),
            "capital_flujo_total":   round(capital_reclas, 2),
        }

    # ── INTERCOMPANY ──
    ic_row = db.execute(text("""
        SELECT COALESCE(SUM(monto_egreso), 0) AS egresado,
               COALESCE(SUM(monto_ingreso), 0) AS ingresado
        FROM flujos_efectivo
        WHERE sociedad ILIKE :s
          AND seccion = 'FINANCIAMIENTO'
          AND LOWER(nombre_categoria) LIKE '%intercompany%'
    """), {"s": f"%{sociedad_flujos}%"}).fetchone()

    saldo_intercompany = float(ic_row.egresado or 0) - float(ic_row.ingresado or 0)
    ic_tipo = "por pagar" if saldo_intercompany > 0 else "a favor"
    ic_abs  = abs(saldo_intercompany)
    ic_dist = round(ic_abs / anos_ic, 2) if anos_ic > 0 else None

    # ── Ejecutado real desde flujos_efectivo vs tabla amortización ──
    fin_real_rows = db.execute(text("""
        SELECT
            nombre_categoria,
            COALESCE(SUM(monto_egreso), 0) - COALESCE(SUM(monto_ingreso), 0) AS neto
        FROM flujos_efectivo
        WHERE sociedad ILIKE :s
          AND seccion = 'FINANCIAMIENTO'
        GROUP BY nombre_categoria
    """), {"s": f"%{sociedad_flujos}%"}).fetchall()

    fin_real_por_categoria = {r.nombre_categoria: float(r.neto) for r in fin_real_rows}
    fin_real_total = sum(fin_real_por_categoria.values())

    # Group pending cuotas by year for chart
    cuotas_por_anio = {}
    if prestamo_result:
        for c in prestamo_result.get("cuotas_pendientes", []):
            yr = c["fecha"][:4]
            if yr not in cuotas_por_anio:
                cuotas_por_anio[yr] = {"capital": 0.0, "interes": 0.0, "cuota": 0.0, "n": 0}
            cuotas_por_anio[yr]["capital"] += c["capital"]
            cuotas_por_anio[yr]["interes"] += c["interes"]
            cuotas_por_anio[yr]["cuota"]   += c["cuota"]
            cuotas_por_anio[yr]["n"]       += 1

    return {
        "empresa": empresa,
        "prestamo_bancario": prestamo_result,
        "intercompany": {
            "saldo": round(ic_abs, 2),
            "tipo":  ic_tipo,
            "anos_distribucion": anos_ic,
            "dist_anual": ic_dist,
        },
        "ejecutado_financiamiento": {
            "por_categoria": fin_real_por_categoria,
            "total": round(fin_real_total, 2),
        },
        "cuotas_por_anio": {
            yr: {k: round(v, 2) if isinstance(v, float) else v for k, v in data.items()}
            for yr, data in sorted(cuotas_por_anio.items())
        },
    }


@router.get("/{empresa}/tierra-dividendos")
async def get_tierra_dividendos(
    empresa: str,
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """Plan de pagos tierra y dividendos: plan total vs pagado vs pendiente."""
    sociedad_flujos = normalizar_sociedad_flujos(empresa)
    hoy = date.today()

    # Pagado real de TERRENO
    tierra_pagado_row = db.execute(text("""
        SELECT COALESCE(SUM(monto_egreso), 0) AS pagado
        FROM flujos_efectivo
        WHERE sociedad ILIKE :s AND seccion = 'TERRENO'
    """), {"s": f"%{sociedad_flujos}%"}).fetchone()
    tierra_pagado = float(tierra_pagado_row.pagado or 0)

    # Pagado real de DIVIDENDOS — desde reclasificaciones (donde se reclasifican intercompany→dividendos)
    # + cualquier pago directo en flujos_efectivo sección DIVIDENDOS
    div_pagado_reclas = db.execute(text("""
        SELECT COALESCE(SUM(monto), 0) AS pagado
        FROM flujos_reclasificaciones
        WHERE sociedad ILIKE :s AND seccion_destino = 'DIVIDENDOS'
    """), {"s": f"%{sociedad_flujos}%"}).fetchone()
    div_pagado_flujo = db.execute(text("""
        SELECT COALESCE(SUM(monto_egreso), 0) + COALESCE(SUM(monto_ingreso), 0) AS pagado
        FROM flujos_efectivo
        WHERE sociedad ILIKE :s AND seccion = 'DIVIDENDOS'
    """), {"s": f"%{sociedad_flujos}%"}).fetchone()
    div_pagado = float(div_pagado_reclas.pagado or 0) + float(div_pagado_flujo.pagado or 0)

    # Plan tierra para esta sociedad
    soc_upper = sociedad_flujos.upper().strip()
    plan_tierra_empresa = []
    for plan in get_plan_tierra():
        plan_soc = plan["sociedad"].upper().strip()
        if plan_soc == soc_upper or plan_soc in soc_upper or soc_upper in plan_soc:
            plan_tierra_empresa.append(plan)

    # Calcular total plan tierra y pendientes
    tierra_plan_total = 0.0
    tierra_pendiente_detalle = []
    for plan in plan_tierra_empresa:
        for fecha_str, monto in plan["pagos"].items():
            tierra_plan_total += monto
            fecha_p = datetime.strptime(fecha_str, "%Y-%m-%d").date()
            if fecha_p > hoy:
                tierra_pendiente_detalle.append({
                    "fecha": fecha_str,
                    "concepto": plan["concepto"],
                    "monto": round(monto, 2)
                })

    tierra_pendiente_detalle.sort(key=lambda x: x["fecha"])
    # tierra_pendiente_total se calcula abajo como plan - pagado real

    # Plan dividendos para esta sociedad
    plan_div_empresa = []
    for plan in get_plan_dividendos():
        plan_soc = plan["sociedad"].upper().strip()
        if plan_soc == soc_upper or plan_soc in soc_upper or soc_upper in plan_soc:
            plan_div_empresa.append(plan)

    div_plan_total = 0.0
    div_pendiente_detalle = []
    for plan in plan_div_empresa:
        for fecha_str, monto in plan["pagos"].items():
            div_plan_total += monto
            fecha_p = datetime.strptime(fecha_str, "%Y-%m-%d").date()
            if fecha_p > hoy:
                div_pendiente_detalle.append({
                    "fecha": fecha_str,
                    "cuenta": plan["cuenta"],
                    "monto": round(monto, 2)
                })

    div_pendiente_detalle.sort(key=lambda x: x["fecha"])
    # Pendiente = Plan Total - Pagado Real (no suma de pagos futuros del plan)
    div_pendiente_total = max(0, div_plan_total - div_pagado)

    # Mismo fix para tierra: pendiente = plan - pagado real
    tierra_pendiente_total = max(0, tierra_plan_total - tierra_pagado)

    return {
        "empresa": empresa,
        "tierra": {
            "plan_total":   round(tierra_plan_total, 2),
            "pagado":       round(tierra_pagado, 2),
            "pendiente":    round(tierra_pendiente_total, 2),
            "avance_pct":   round(tierra_pagado / tierra_plan_total * 100, 1) if tierra_plan_total else 0,
            "proximos_pagos": tierra_pendiente_detalle[:24],
        },
        "dividendos": {
            "plan_total":   round(div_plan_total, 2),
            "pagado":       round(div_pagado, 2),
            "pendiente":    round(div_pendiente_total, 2),
            "avance_pct":   round(div_pagado / div_plan_total * 100, 1) if div_plan_total else 0,
            "proximos_pagos": div_pendiente_detalle[:24],
        },
    }


@router.get("/{empresa}/flujo")
async def get_flujo(empresa: str, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Flujo proyectado al cierre — consolida todos los módulos."""
    sup = db.execute(
        text("SELECT * FROM proyeccion_supuestos WHERE empresa = :e LIMIT 1"),
        {"e": empresa}
    ).fetchone()
    if not sup:
        raise HTTPException(404,
            "No hay supuestos guardados. Completá los datos y presioná 'Guardar cambios'.")

    anos = int(sup.anos_proyecto or 5)
    tasa_desc = float(sup.tasa_descuento or 0.12)
    pct_isr = float(sup.pct_isr or 0)
    sociedad_flujos = normalizar_sociedad_flujos(empresa)
    anos_ic_saved = int(sup.anos_ic or 0) if hasattr(sup, 'anos_ic') and sup.anos_ic else 0
    anos_egr_saved = int(sup.anos_egr or 0) if hasattr(sup, 'anos_egr') and sup.anos_egr else 0
    _eo_raw = sup.egresos_operativos
    if isinstance(_eo_raw, str):
        try: plazos = json.loads(_eo_raw)
        except: plazos = []
    elif isinstance(_eo_raw, list):
        plazos = _eo_raw
    else:
        plazos = [] if _eo_raw is None else list(_eo_raw)

    # Llamar sub-endpoints con keyword args y los años guardados
    ing  = await get_ingresos(empresa=empresa, db=db, user=user)
    egr  = await get_egresos_operativos(empresa=empresa, anos_egr=anos_egr_saved, db=db, user=user)
    fin  = await get_egresos_financieros(empresa=empresa, anos_ic=anos_ic_saved, db=db, user=user)
    tier = await get_tierra_dividendos(empresa=empresa, db=db, user=user)

    # Calcular ingresos proyectados desde plazos
    total_ing_proyectado = 0.0
    for p in plazos:
        ini_str = p.get("inicio", "")
        fin_str = p.get("fin", "")
        tk      = float(p.get("ticket", 0))
        tasa_p  = float(p.get("tasa", 0))
        plazo_m = int(p.get("plazo", 0)) if str(p.get("plazo","")).isdigit() else 0
        unid_mes= int(p.get("unidMes", 0))
        if not ini_str or not fin_str or not tk or not unid_mes: continue
        interes_u = tk * (tasa_p/100) * (plazo_m/12) if plazo_m > 0 else 0
        total_ing_proyectado += (tk + interes_u) * int(p.get("unidProy", unid_mes))

    total_ing_real = ing["ingresos_reales"]["saldo_pendiente"]
    total_ing = total_ing_real + total_ing_proyectado

    # IVA
    cap_real = ing["ingresos_reales"]["saldo_capital"]
    cap_proy = sum(float(p.get("ticket", 0)) * int(p.get("unidProy", int(p.get("unidMes",0)))) for p in plazos)
    iva_deb  = (cap_real + cap_proy) * 0.70 * 0.12
    egr_op_total = egr["total"]["pendiente"]
    iva_cred = egr_op_total * 0.12
    iva_neto = iva_deb - iva_cred

    # Financiero
    prest = fin.get("prestamo_bancario") or {}
    egr_prest = prest.get("pendiente_total", 0)
    _ic_saldo = fin["intercompany"]["saldo"]
    _ic_tipo  = fin["intercompany"].get("tipo", "")
    # Usar el campo tipo calculado en /financieros (saldo>0="por pagar", saldo<0="a favor")
    # "por pagar" → empresa debe pagar → egreso (positivo con paréntesis)
    # "a favor"   → empresa va a cobrar → ingreso (negativo sin paréntesis)
    if _ic_tipo == "por pagar":
        egr_ic = abs(_ic_saldo)   # positivo = egreso
    elif _ic_tipo == "a favor":
        egr_ic = -abs(_ic_saldo)  # negativo = ingreso
    else:
        egr_ic = -_ic_saldo       # fallback: invertir signo
    egr_fin = egr_prest + egr_ic

    # Tierra y dividendos
    egr_tierra = tier["tierra"]["pendiente"]
    egr_div = tier["dividendos"]["pendiente"]

    # ISR
    isr = total_ing * pct_isr

    flujo_neto = total_ing - egr_op_total - iva_neto - egr_fin - isr - egr_tierra - egr_div

    # ── Distribución anual de ingresos reales por fecha_programada_cobro ──
    anio_actual = date.today().year
    ing_real_rows = db.execute(text("""
        SELECT EXTRACT(YEAR FROM fecha_programada_cobro) AS anio,
               SUM(saldo_pendiente) AS total
        FROM ov_cartera
        WHERE empresa = :e AND line_status = 'O'
          AND fecha_programada_cobro IS NOT NULL
        GROUP BY anio ORDER BY anio
    """), {"e": normalizar_empresa_cartera(empresa)}).fetchall()
    ing_real_por_anio = {int(r.anio): float(r.total) for r in ing_real_rows}

    # ── Distribución anual de ingresos proyectados por plazos ──
    # Cada venta genera cobros mensuales a lo largo del plazo.
    # Ejemplo: venta en jun-2027 a 60 meses → cobros jun-2027 a jun-2032
    # Contado (plazo=0) se cobra en el mes de la venta.
    ing_proy_por_anio = {}
    for p in plazos:
        ini_str = p.get("inicio", "")
        fin_str = p.get("fin", "")
        tk      = float(p.get("ticket", 0))
        tasa_p  = float(p.get("tasa", 0))
        plazo_m = int(p.get("plazo", 0)) if str(p.get("plazo","")).isdigit() else 0
        unid_mes= int(p.get("unidMes", 0))
        if not ini_str or not fin_str or not tk or not unid_mes: continue
        try:
            ini_d = datetime.strptime(ini_str + "-01", "%Y-%m-%d").date()
            fin_d = datetime.strptime(fin_str + "-01", "%Y-%m-%d").date()
        except: continue
        interes_u = tk * (tasa_p/100) * (plazo_m/12) if plazo_m > 0 else 0
        ingreso_total_u = tk + interes_u  # total por unidad (capital + intereses)

        # Recorrer cada mes de ventas (desde inicio hasta fin)
        cur_venta = ini_d
        while cur_venta < fin_d:
            # En este mes se venden unid_mes unidades
            if plazo_m <= 0:
                # Contado: se cobra todo en el mes de venta
                yr_cobro = cur_venta.year
                ing_proy_por_anio[yr_cobro] = ing_proy_por_anio.get(yr_cobro, 0) + ingreso_total_u * unid_mes
            else:
                # Crédito: distribuir cobro mensual a lo largo del plazo
                cuota_mensual = ingreso_total_u / plazo_m  # cuota mensual por unidad
                total_cuota_mes = cuota_mensual * unid_mes  # total de cuotas de las unid vendidas este mes
                cur_cobro = cur_venta
                for _ in range(plazo_m):
                    yr_cobro = cur_cobro.year
                    ing_proy_por_anio[yr_cobro] = ing_proy_por_anio.get(yr_cobro, 0) + total_cuota_mes
                    # Avanzar un mes en los cobros
                    m_c = cur_cobro.month + 1
                    y_c = cur_cobro.year + (1 if m_c > 12 else 0)
                    m_c = m_c if m_c <= 12 else 1
                    try: cur_cobro = cur_cobro.replace(year=y_c, month=m_c)
                    except: break
            # Avanzar un mes en las ventas
            mes_v = cur_venta.month + 1
            anio_v = cur_venta.year + (1 if mes_v > 12 else 0)
            mes_v = mes_v if mes_v <= 12 else 1
            try: cur_venta = cur_venta.replace(year=anio_v, month=mes_v)
            except: break

    # Determinar rango real de años
    todos_anios = set(ing_real_por_anio.keys()) | set(ing_proy_por_anio.keys())
    if todos_anios:
        anio_max = max(todos_anios)
        anos_real = max(1, anio_max - anio_actual + 1)
    else:
        anos_real = anos
    anos_total = max(anos_real, anos)

    # Egresos operativos distribuidos según años configurados (anos_egr o anos si no hay)
    anos_egr_eff = anos_egr_saved if anos_egr_saved > 0 else anos
    anos_ic_eff  = anos_ic_saved if anos_ic_saved > 0 else anos
    egr_a  = egr_op_total / anos_egr_eff if anos_egr_eff else egr_op_total
    # iva_a/isr_a removed — computed per year in loop

    # Cuotas préstamo por año calendario
    cuotas_pend = prest.get("cuotas_pendientes", [])
    hoy = date.today()
    fin_yr = {}
    for c in cuotas_pend:
        try:
            yr = int(c["fecha"][:4]) - anio_actual + 1
            yr = max(1, yr)
            fin_yr[yr] = fin_yr.get(yr, 0) + c["cuota"]
        except: pass
    fin_yr_ic = egr_ic / anos_ic_eff if anos_ic_eff else egr_ic

    # Tierra y dividendos: distribuir por plan (fecha de cada pago pendiente → año calendario)
    def _dist_plan_scaled(plan_list, sociedad_flujos, hoy, anio_actual, pendiente_total):
        raw = {}
        for plan in plan_list:
            plan_soc = plan["sociedad"].upper().strip()
            if plan_soc not in sociedad_flujos.upper() and sociedad_flujos.upper() not in plan_soc:
                continue
            for f_str, monto in plan["pagos"].items():
                try:
                    f_d = datetime.strptime(f_str, "%Y-%m-%d").date()
                    if f_d > hoy:
                        yr_k = f_d.year - anio_actual + 1
                        yr_k = max(1, yr_k)
                        raw[yr_k] = raw.get(yr_k, 0) + monto
                except: pass
        raw_total = sum(raw.values())
        if raw_total <= 0 or pendiente_total <= 0:
            return raw
        factor = pendiente_total / raw_total
        return {yr: round(v * factor, 2) for yr, v in raw.items()}

    tier_por_anio = _dist_plan_scaled(get_plan_tierra(), sociedad_flujos, hoy, anio_actual, egr_tierra)
    div_por_anio  = _dist_plan_scaled(get_plan_dividendos(), sociedad_flujos, hoy, anio_actual, egr_div)

    # ── Desglose de cuotas préstamo por año: capital vs intereses ──
    prest_cap_por_anio = {}
    prest_int_por_anio = {}
    cuotas_pend_raw = prest.get("cuotas_pendientes", []) or []
    for c in cuotas_pend_raw:
        try:
            yr_c = int(c["fecha"][:4]) - anio_actual + 1
            yr_c = max(1, yr_c)
            prest_cap_por_anio[yr_c] = prest_cap_por_anio.get(yr_c, 0) + c["capital"]
            prest_int_por_anio[yr_c] = prest_int_por_anio.get(yr_c, 0) + c["interes"]
        except: pass

    # ── Desglose egresos operativos: urbanización vs administración ──
    egr_urb = egr.get("urbanizacion", {})
    egr_adm = egr.get("administracion", {})
    egr_urb_pend = float(egr_urb.get("pendiente", 0) if isinstance(egr_urb, dict) else 0)
    egr_adm_pend = float(egr_adm.get("pendiente", 0) if isinstance(egr_adm, dict) else 0)
    egr_urb_a = egr_urb_pend / anos_egr_eff if anos_egr_eff else 0
    egr_adm_a = egr_adm_pend / anos_egr_eff if anos_egr_eff else 0

    flujo_anual = []
    pico = {"anio": None, "flujo": 0.0}
    for yr in range(1, anos_total + 1):
        anio_cal = anio_actual + yr - 1
        ing_real_yr = ing_real_por_anio.get(anio_cal, 0)
        ing_proy_yr = ing_proy_por_anio.get(anio_cal, 0)
        ing_yr = ing_real_yr + ing_proy_yr

        # Egresos operativos desglosados
        urb_yr = egr_urb_a if yr <= anos_egr_eff else 0
        adm_yr = egr_adm_a if yr <= anos_egr_eff else 0
        egr_yr = urb_yr + adm_yr

        if yr <= anos:
            iva_deb_yr  = ing_yr * 0.70 * 0.12
            iva_cred_yr = (urb_yr + adm_yr) * 0.12
            iva_yr = max(0, iva_deb_yr - iva_cred_yr)
            isr_yr = ing_yr * pct_isr
        else:
            iva_yr = 0.0
            isr_yr = 0.0
        tier_yr = tier_por_anio.get(yr, 0)
        div_yr  = div_por_anio.get(yr, 0)

        # Egresos financieros desglosados
        prest_cap_yr = prest_cap_por_anio.get(yr, 0)
        prest_int_yr = prest_int_por_anio.get(yr, 0)
        ic_yr = fin_yr_ic if yr <= anos_ic_eff else 0
        efin_yr = prest_cap_yr + prest_int_yr + ic_yr

        fno_yr = ing_yr - egr_yr   # Flujo Neto de Operaciones (ingresos - egresos op)
        fn_yr = ing_yr - egr_yr - iva_yr - efin_yr - isr_yr - tier_yr - div_yr
        acum  = (flujo_anual[-1]["flujo_acumulado"] if flujo_anual else 0) + fn_yr
        flujo_anual.append({
            "anio": yr, "anio_cal": anio_cal,
            "ingresos":          round(ing_yr, 2),
            "ing_real":          round(ing_real_yr, 2),
            "ing_proy":          round(ing_proy_yr, 2),
            "urbanizacion":      round(urb_yr, 2),
            "urb_proy":          round(urb_yr, 2),
            "administracion":    round(adm_yr, 2),
            "adm_proy":          round(adm_yr, 2),
            "egresos_op":        round(egr_yr, 2),
            "intercompany":      round(ic_yr, 2),
            "ic_proy":           round(ic_yr, 2),
            "prestamo_capital":  round(prest_cap_yr, 2),
            "prest_cap_proy":    round(prest_cap_yr, 2),
            "prestamo_interes":  round(prest_int_yr, 2),
            "prest_int_proy":    round(prest_int_yr, 2),
            "egresos_fin":       round(efin_yr, 2),
            "iva_neto":          round(iva_yr, 2),
            "iva_proy":          round(iva_yr, 2),
            "isr":               round(isr_yr, 2),
            "isr_proy":          round(isr_yr, 2),
            "tierra":            round(tier_yr, 2),
            "tierra_proy":       round(tier_yr, 2),
            "dividendos":        round(div_yr, 2),
            "div_proy":          round(div_yr, 2),
            "tierra_capital":    round(tier_yr + div_yr, 2),
            "flujo_neto_op":     round(fno_yr, 2),
            "flujo_neto":        round(fn_yr, 2),
            "flujo_acumulado":   round(acum, 2),
            "es_negativo":       fn_yr < 0,
        })
        if fn_yr < pico["flujo"]:
            pico = {"anio": yr, "flujo": round(fn_yr, 2)}

    inv0 = -(egr_tierra + egr_div + prest.get("pendiente_capital", 0))
    flujos_calc = [inv0] + [f["flujo_neto_op"] for f in flujo_anual]
    tir = calcular_tir(flujos_calc)
    van = calcular_van(flujos_calc, tasa_desc)

    u_total = ing["total_disponibles"] + ing["ingresos_reales"]["contratos"]
    ticket_prom = total_ing / u_total if u_total else 0
    lotes_pe = int((egr_op_total + iva_neto + egr_fin + isr + egr_tierra + egr_div) / ticket_prom) if ticket_prom > 0 else 0
    margen = round(flujo_neto / total_ing * 100, 2) if total_ing else 0

    # ── Saldo inicial desde flujos_saldo_inicial ──────────────────────────
    saldo_ini_row = db.execute(text("""
        SELECT monto FROM flujos_saldo_inicial
        WHERE sociedad ILIKE :s ORDER BY id ASC LIMIT 1
    """), {"s": f"%{sociedad_flujos}%"}).fetchone()
    saldo_inicial_real = float(saldo_ini_row.monto if saldo_ini_row else 0)

    # ── Flujo real con reclasificaciones aplicadas (cuadra con sección Flujos) ──
    flujo_real_rows = db.execute(text("""
        WITH base AS (
            SELECT anio,
                COALESCE(SUM(CASE WHEN seccion='INGRESOS'
                    THEN monto_ingreso - monto_egreso ELSE 0 END),0)                         AS ingresos,
                COALESCE(SUM(CASE WHEN seccion='EGRESOS / URBANIZACION'
                    THEN monto_egreso - monto_ingreso ELSE 0 END),0)                         AS urb,
                COALESCE(SUM(CASE WHEN seccion IN (
                    'EGRESOS / MOVIMIENTO DE TIERRAS','EGRESOS / MOV. TIERRAS / MAQUINARIA')
                    THEN monto_egreso - monto_ingreso ELSE 0 END),0)                         AS mov,
                COALESCE(SUM(CASE WHEN seccion='EGRESOS / ADMINISTRACION'
                    THEN monto_egreso - monto_ingreso ELSE 0 END),0)                         AS adm,
                COALESCE(SUM(CASE WHEN seccion='FINANCIAMIENTO'
                    THEN monto_egreso - monto_ingreso ELSE 0 END),0)                         AS fin,
                COALESCE(SUM(CASE WHEN seccion='IMPUESTOS'
                    THEN monto_egreso - monto_ingreso ELSE 0 END),0)                         AS impuestos,
                COALESCE(SUM(CASE WHEN seccion='TERRENO'
                    THEN monto_egreso - monto_ingreso ELSE 0 END),0)                         AS tierra
            FROM flujos_efectivo WHERE sociedad ILIKE :s GROUP BY anio
        ),
        adj AS (
            SELECT EXTRACT(YEAR FROM fecha_contable)::int AS anio,
                COALESCE(SUM(CASE WHEN seccion_destino='EGRESOS / URBANIZACION'   AND seccion_origen!=seccion_destino THEN monto ELSE 0 END),0)
               -COALESCE(SUM(CASE WHEN seccion_origen ='EGRESOS / URBANIZACION'   AND seccion_origen!=seccion_destino THEN monto ELSE 0 END),0) AS urb_adj,
                COALESCE(SUM(CASE WHEN seccion_destino='EGRESOS / ADMINISTRACION' AND seccion_origen!=seccion_destino THEN monto ELSE 0 END),0)
               -COALESCE(SUM(CASE WHEN seccion_origen ='EGRESOS / ADMINISTRACION' AND seccion_origen!=seccion_destino THEN monto ELSE 0 END),0) AS adm_adj,
                COALESCE(SUM(CASE WHEN seccion_destino='FINANCIAMIENTO'           AND seccion_origen!=seccion_destino THEN monto ELSE 0 END),0)
               -COALESCE(SUM(CASE WHEN seccion_origen ='FINANCIAMIENTO'           AND seccion_origen!=seccion_destino THEN monto ELSE 0 END),0) AS fin_adj
            FROM flujos_reclasificaciones WHERE sociedad ILIKE :s GROUP BY 1
        )
        SELECT b.anio AS anio_cal,
               b.ingresos,
               b.urb + COALESCE(a.urb_adj,0) AS urbanizacion,
               b.mov                          AS mov_tierras,
               b.adm + COALESCE(a.adm_adj,0) AS administracion,
               b.fin + COALESCE(a.fin_adj,0)  AS financiamiento,
               b.impuestos,
               b.tierra
        FROM base b LEFT JOIN adj a ON a.anio=b.anio ORDER BY b.anio
    """), {"s": f"%{sociedad_flujos}%"}).fetchall()

    # Intereses bancarios reales desde reclasificaciones
    intereses_reclas_rows = db.execute(text("""
        SELECT EXTRACT(YEAR FROM fecha_contable)::int AS anio, SUM(monto) AS total
        FROM flujos_reclasificaciones
        WHERE sociedad ILIKE :s AND seccion_destino='FINANCIAMIENTO'
          AND (LOWER(nombre_destino) LIKE '%interes%' OR LOWER(nombre_destino) LIKE '%interés%')
        GROUP BY 1
    """), {"s": f"%{sociedad_flujos}%"}).fetchall()
    intereses_reclas_por_anio = {int(r.anio): float(r.total or 0) for r in intereses_reclas_rows}

    # Intercompany real RAW por año (sin restar IC→DIV, eso ya está en financiamiento_neto)
    ic_real_rows = db.execute(text("""
        SELECT EXTRACT(YEAR FROM fecha_contable)::int AS anio,
               COALESCE(SUM(monto_egreso - monto_ingreso),0) AS ic_neto
        FROM flujos_efectivo WHERE sociedad ILIKE :s
          AND seccion='FINANCIAMIENTO' AND LOWER(nombre_categoria) LIKE '%intercompany%'
        GROUP BY 1
    """), {"s": f"%{sociedad_flujos}%"}).fetchall()
    ic_real_por_anio = {int(r.anio): float(r.ic_neto or 0) for r in ic_real_rows}

    # Préstamo bancario real neto por año
    prest_real_rows = db.execute(text("""
        SELECT EXTRACT(YEAR FROM fecha_contable)::int AS anio,
               COALESCE(SUM(monto_egreso - monto_ingreso),0) AS prest_neto
        FROM flujos_efectivo WHERE sociedad ILIKE :s
          AND seccion='FINANCIAMIENTO' AND LOWER(nombre_categoria) LIKE '%prestamo%'
        GROUP BY 1
    """), {"s": f"%{sociedad_flujos}%"}).fetchall()
    prest_real_por_anio = {int(r.anio): float(r.prest_neto or 0) for r in prest_real_rows}

    div_real_rows = db.execute(text("""
        SELECT EXTRACT(YEAR FROM fecha_contable)::int AS anio, SUM(monto) AS total
        FROM flujos_reclasificaciones
        WHERE sociedad ILIKE :s AND seccion_destino = 'DIVIDENDOS'
        GROUP BY 1
    """), {"s": f"%{sociedad_flujos}%"}).fetchall()
    div_real_por_anio = {int(r.anio): float(r.total or 0) for r in div_real_rows}

    flujo_real_anual = []
    acum_real = saldo_inicial_real
    for r in flujo_real_rows:
        yr_cal = int(r.anio_cal)
        if yr_cal > anio_actual: continue
        ing_r  = float(r.ingresos       or 0)
        urb_r  = float(r.urbanizacion   or 0)
        mov_r  = float(r.mov_tierras    or 0)
        adm_r  = float(r.administracion or 0)
        fin_r  = float(r.financiamiento or 0)
        pint_r_reclas = intereses_reclas_por_anio.get(yr_cal, 0)
        imp_r  = float(r.impuestos or 0)
        tier_r = float(r.tierra    or 0)
        div_r  = div_real_por_anio.get(yr_cal, 0)
        egr_r  = urb_r + mov_r + adm_r
        fn_r   = ing_r - egr_r - fin_r - imp_r - tier_r - div_r
        acum_real += fn_r
        es_anio_mixto = (yr_cal == anio_actual)
        flujo_real_anual.append({
            "anio_cal":            yr_cal,
            "es_mixto":            es_anio_mixto,
            "ingresos":            round(ing_r, 2),
            "ing_real":            round(ing_r, 2),
            "ing_proy":            0,
            "urbanizacion":        round(urb_r, 2),
            "mov_tierras":         round(mov_r, 2),
            "administracion":      round(adm_r, 2),
            "egresos_op":          round(egr_r, 2),
            "financiamiento_neto": round(fin_r, 2),
            "prestamo_interes":    round(pint_r_reclas, 2),
            "intercompany":        round(ic_real_por_anio.get(yr_cal, 0), 2),
            "prestamo_capital":    round(prest_real_por_anio.get(yr_cal, 0), 2),
            "otros_fin":           0,
            "egresos_fin":         round(fin_r, 2),
            "iva_neto":            round(imp_r, 2),
            "isr":                 0,
            "tierra":              round(tier_r, 2),
            "dividendos":          round(div_r, 2),
            "tierra_capital":      round(tier_r + div_r, 2),
            "flujo_neto":          round(fn_r, 2),
            "flujo_acumulado":     round(acum_real, 2),
        })

    # Ajustar acumulado de proyección para que continúe desde el real
    if flujo_real_anual and flujo_anual:
        base_acum = flujo_real_anual[-1]["flujo_acumulado"]
        for fa in flujo_anual:
            fa["flujo_acumulado"] = round(fa["flujo_acumulado"] + base_acum, 2)

    return {
        "empresa": empresa,
        "saldo_inicial_real": round(saldo_inicial_real, 2),
        "resumen_ingresos": {"real": total_ing_real, "proyectado": total_ing_proyectado, "total": total_ing},
        "resumen_egresos_op": egr["total"],
        "iva": {"debito": round(iva_deb,2), "credito": round(iva_cred,2), "neto": round(iva_neto,2)},
        "egresos_financieros": {"prestamo": egr_prest, "intercompany": egr_ic, "total": round(egr_fin,2)},
        "isr": {"pct": pct_isr, "total": round(isr, 2)},
        "tierra_dividendos": {"tierra": egr_tierra, "dividendos": egr_div, "total": round(egr_tierra+egr_div,2)},
        "flujo_neto_total": round(flujo_neto, 2),
        "flujo_anual": flujo_anual,
        "flujo_real_anual": flujo_real_anual,
        "indicadores": {
            "tir": tir, "van": round(van,2), "margen_neto": margen,
            "punto_equilibrio_lotes": lotes_pe,
            "pico_negativo": pico if pico["anio"] else None,
        },
    }

@router.delete("/{empresa}")
async def delete_supuestos(empresa: str, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Elimina los supuestos guardados de un proyecto para empezar de cero."""
    db.execute(text("DELETE FROM proyeccion_supuestos WHERE empresa = :e"), {"e": empresa})
    db.commit()
    return {"ok": True, "empresa": empresa, "mensaje": "Supuestos eliminados correctamente"}

@router.get("/cxp")
async def get_cuentas_por_pagar(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """
    Cuentas por pagar consolidadas — datos completos sin filtro de fecha.
    Visualización histórica completa para Junta Directiva.
    """
    try:
        rows = db.execute(text("""
            SELECT empresa, doc_num, codigo_proveedor, nombre_acreedor,
                   ref_acreedor, fecha_vencimiento, importe, comentarios
            FROM cuentas_por_pagar
            ORDER BY empresa, fecha_vencimiento
        """)).fetchall()

        items = []
        total = 0.0
        for r in rows:
            imp = float(r.importe or 0)
            total += imp
            items.append({
                "empresa": r.empresa,
                "doc_num": str(r.doc_num) if r.doc_num else None,
                "codigo_proveedor": r.codigo_proveedor,
                "nombre_acreedor": r.nombre_acreedor,
                "ref_acreedor": r.ref_acreedor,
                "fecha_vencimiento": str(r.fecha_vencimiento)[:10] if r.fecha_vencimiento else None,
                "importe": imp,
                "comentarios": r.comentarios,
            })

        # Resumen por empresa (para totales en frontend)
        resumen_empresa = {}
        for it in items:
            e = it["empresa"] or "—"
            resumen_empresa.setdefault(e, {"empresa": e, "total": 0.0, "count": 0})
            resumen_empresa[e]["total"] += it["importe"]
            resumen_empresa[e]["count"] += 1

        return {
            "items": items,
            "total_general": total,
            "total_registros": len(items),
            "por_empresa": sorted(resumen_empresa.values(), key=lambda x: -x["total"]),
        }
    except Exception as e:
        # Si la tabla no existe (sync nunca ejecutado) devolver vacío sin error
        return {"items": [], "total_general": 0, "total_registros": 0, "por_empresa": [], "_warning": str(e)[:200]}


@router.post("/cxp/sync")
async def sync_cuentas_por_pagar(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    """
    Sincroniza la tabla cuentas_por_pagar desde la pestaña 'CUENTA POR PAGAR'
    del Excel 'PRESUPUESTO, PRESTAMOS Y TIERRA.xlsx'.
    El usuario sube el archivo, el endpoint procesa y reemplaza los registros.
    """
    # Validar archivo
    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(status_code=400, detail="Archivo debe ser .xlsx o .xlsm")

    try:
        import pandas as pd
    except ImportError:
        raise HTTPException(status_code=500, detail="pandas no instalado en el backend")

    content = await file.read()
    try:
        xl = pd.ExcelFile(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo abrir el Excel: {e}")

    if 'CUENTA POR PAGAR' not in xl.sheet_names:
        raise HTTPException(
            status_code=400,
            detail=f"Pestaña 'CUENTA POR PAGAR' no encontrada. Pestañas disponibles: {xl.sheet_names}"
        )

    df = pd.read_excel(xl, sheet_name='CUENTA POR PAGAR', header=0)
    df.columns = [str(c).strip() for c in df.columns]

    # Validar columnas requeridas
    required = ['EMPRESA', 'Nº documento', 'Código de proveedor', 'Nombre de acreedor',
                'No.Ref.del acreedor', 'Fecha de vencimiento', 'Importe', 'Comentarios']
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Columnas faltantes en la pestaña: {missing}. Encontradas: {list(df.columns)}"
        )

    # Filtrar filas vacías
    df = df.dropna(subset=['EMPRESA', 'Importe'], how='all')
    df = df[df['EMPRESA'].astype(str).str.strip() != '']

    # Asegurar que la tabla existe
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS cuentas_por_pagar (
            id SERIAL PRIMARY KEY,
            empresa VARCHAR(150),
            doc_num VARCHAR(50),
            codigo_proveedor VARCHAR(50),
            nombre_acreedor VARCHAR(300),
            ref_acreedor VARCHAR(150),
            fecha_vencimiento DATE,
            importe NUMERIC(14,2),
            comentarios TEXT,
            sincronizado_en TIMESTAMP DEFAULT NOW()
        )
    """))

    # Limpiar tabla antes de re-cargar (full refresh)
    db.execute(text("TRUNCATE TABLE cuentas_por_pagar RESTART IDENTITY"))

    # Insertar
    insertados = 0
    errores = []
    for idx, row in df.iterrows():
        try:
            fv = row['Fecha de vencimiento']
            if pd.isna(fv):
                fv_str = None
            elif hasattr(fv, 'date'):
                fv_str = fv.date()
            else:
                fv_str = pd.to_datetime(fv).date()

            db.execute(text("""
                INSERT INTO cuentas_por_pagar
                  (empresa, doc_num, codigo_proveedor, nombre_acreedor,
                   ref_acreedor, fecha_vencimiento, importe, comentarios)
                VALUES
                  (:e, :dn, :cp, :na, :ra, :fv, :imp, :com)
            """), {
                "e":   str(row['EMPRESA']).strip()[:150] if pd.notna(row['EMPRESA']) else None,
                "dn":  str(row['Nº documento']).strip()[:50] if pd.notna(row['Nº documento']) else None,
                "cp":  str(row['Código de proveedor']).strip()[:50] if pd.notna(row['Código de proveedor']) else None,
                "na":  str(row['Nombre de acreedor']).strip()[:300] if pd.notna(row['Nombre de acreedor']) else None,
                "ra":  str(row['No.Ref.del acreedor']).strip()[:150] if pd.notna(row['No.Ref.del acreedor']) else None,
                "fv":  fv_str,
                "imp": float(row['Importe']) if pd.notna(row['Importe']) else 0,
                "com": str(row['Comentarios']).strip() if pd.notna(row['Comentarios']) else None,
            })
            insertados += 1
        except Exception as e:
            errores.append({"fila": int(idx) + 2, "error": str(e)[:200]})

    db.commit()

    return {
        "ok": True,
        "insertados": insertados,
        "errores": errores[:20],  # primeros 20 errores si hay
        "total_errores": len(errores),
        "mensaje": f"Sincronizados {insertados} registros de cuentas por pagar"
    }
